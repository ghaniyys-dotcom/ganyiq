#!/usr/bin/env python3
"""UNKNOWN ELIMINATION — Deep dives on top 5 unknowns"""

import openpyxl
from openpyxl.styles import Font, PatternFill

PATH = '/root/GANYIQ/OPUS-AUDIT/opus_audit_master.xlsx'
wb = openpyxl.load_workbook(PATH)

HEADER_FILL = PatternFill(start_color='1F1F2E', end_color='1F1F2E', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
ALT_FILL = PatternFill(start_color='F5F5FF', end_color='F5F5FF', fill_type='solid')

def style_header(ws, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

def write_sheet(name, headers, rows):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(title=name)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=str(val))
        if r % 2 == 0:
            for c in range(1, len(headers)+1):
                ws.cell(row=r, column=c).fill = ALT_FILL
    style_header(ws, len(headers))
    ws.auto_filter.ref = f"A1:{chr(64+len(headers))}{len(rows)+1}"
    print(f"  {name}: {len(rows)} rows")

# ===================================================================
# CANDIDATE_GENERATION_DEEP_DIVE
# ===================================================================
CANDIDATE_GENERATION = [
    ["Claim", "Evidence", "Source", "Confidence", "Impact on Rewrite Decision"],
    ("Scene/timeline based candidate generation", "Bundle strings: 'timeline', 'scene selection', 'Arrange Scenes in Timeline', 'drag the start or end of segments on timeline'", "/tmp/app.js literal strings", "HIGH", "MEDIUM — Confirms user can see and rearrange candidate segments on a timeline. Implies scene-detection-based chunking."),
    ("Editing script with tracks", "editingScript?.tracks — '?include=editingScript' returns editing script with tracks property", "/tmp/app.js exportable-clips API", "HIGH", "HIGH — Tracks-based editing script confirms NLE approach. Each candidate clip has a multi-track editing script."),
    ("B-roll insertion at timeline positions", "POST /brolls with {selectedTimeline, editedTimeline, timelineIn: 1e4}. captionTrack separate field.", "/tmp/app.js brolls API", "HIGH", "MEDIUM — Users can insert b-roll at specific timeline positions. Timeline positions in milliseconds."),
    ("Speaker-turn based candidates", "SpeakerActivityTracker + turn detection in DecisionEngine. Speaker switches trigger peak detection.", "GANYIQ port + Bundle patterns", "MEDIUM", "MEDIUM — Speaker turns likely used as candidate boundaries but not confirmed as primary chunking strategy."),
    ("Hybrid approach (most likely)", "scene detection + speaker turns + transcript paragraphs + emotion peaks combined", "Multiple evidence points converge on hybrid strategy", "MEDIUM", "HIGH — The most likely approach combines multiple signals for candidate boundary detection."),
    ("Manual reframing as fallback", "'Manual Reframing' feature: 'face detection' + 'select scenes' → user can manually reframe. jobType:\"reframe\" exists.", "/tmp/app.js literal + help.opus.pro docs", "HIGH", "LOW — Manual reframing is UX feature, not core candidate generation."),
    ("Clip durations configurable", "clipDurations array in curationPref. Users set preferred clip lengths.", "Bundle + official docs", "HIGH", "LOW — Duration preference affects candidate filtering, not generation."),
    ("Topic keywords affect selection", "topicKeywords in curationPref. Official docs: 'Ensure keywords are present in speech or transcript'.", "Bundle + official docs", "HIGH", "LOW — Keywords filter candidates, don't generate them."),
    ("TimelineIn positions exact", "timelineIn: 1e4 (10 seconds). Timeline positions measured in milliseconds.", "/tmp/app.js brolls payload", "HIGH", "LOW — Confirms millisecond-precision timeline."),
]

# ===================================================================
# SCORING_ENGINE_DEEP_DIVE
# ===================================================================
SCORING_ENGINE = [
    ["Claim", "Evidence", "Source", "Confidence", "Impact on Rewrite Decision"],
    ("3 confirmed scoring dimensions", "Official docs: 'Hook: Does the introduction grab attention?' + 'Flow: Does the video flow logically?' + 'Value: Does the video offer valuable knowledge?'", "help.opus.pro Virality Score page", "HIGH", "HIGH — Core scoring framework is confirmed. 0-99 score range confirmed."),
    ("4th Trend dimension (UI only?)", "JS bundle: trendScore + trendComment keys in scoring UI. NOT in official docs.", "/tmp/app.js UI config", "MEDIUM", "MEDIUM — Trend may be part of ClipAnything relevance scoring, not core virality score."),
    ("judgeResult with sub-scores", "judgeResult:{isCopilotClip,relevanceScore,trendTopic,score}. trendTopic extracted as keys.", "/tmp/app.js clip data model", "HIGH", "HIGH — Judge result object provides additional scoring metadata."),
    ("Feedback endpoint: like/dislike", "PUT /curated-clips/{id}.{archId}/feedback. like/dislike fields on clips. f.dislike, f.isEdited tracked.", "/tmp/app.js literal + feedback endpoint", "HIGH", "HIGH — Users can provide explicit feedback on clips. This feeds the scoring system."),
    ("Score range 0-99", "Official docs: 'Virality score ranges from 0 to 99, with higher scores suggesting greater chance of engagement.'", "help.opus.pro Virality Score page", "HIGH", "LOW — Score range confirmed but doesn't reveal calculation method."),
    ("Default sort by virality score", "Official docs: 'By default, your clips will be sorted by Virality Score, from highest to lowest.'", "help.opus.pro Virality Score page", "HIGH", "LOW — Sort behavior confirmed."),
    ("Prompt relevance modifier (ClipAnything)", "Official docs: 'evaluates whether the clip is relevant to your prompt, when using ClipAnything model.'", "help.opus.pro Virality Score page", "HIGH", "MEDIUM — Prompt relevance is a modifier, not a separate dimension."),
    ("LLM vs smaller ML model undetermined", "Docs say 'AI evaluates'. Bundle has 0 LLM API references. Could be LLM or BERT-classifier.", "No backend evidence", "LOW", "HIGH — Critical for GANYIQ strategy. If LLM: need API access. If ML: need training data."),
    ("Scoring is AI-based, not rule-based", "Official docs: 'OpusClip's AI evaluates multiple aspects of your video to determine its Virality Score.'", "help.opus.pro Virality Score page", "HIGH", "MEDIUM — Confirms AI/ML approach over rule-based. But 'AI' ≠ 'LLM'."),
]

# ===================================================================
# RANKING_ENGINE_DEEP_DIVE
# ===================================================================
RANKING_ENGINE = [
    ["Claim", "Evidence", "Source", "Confidence", "Impact on Rewrite Decision"],
    ("Sort: rank > score > judgeResult.score", "Sort function literal: 'if(e.rank&&t.rank)return e.rank-t.rank;let o=e.score||0,i=t.score||0;return o!==i?i-o:((null===(n=t.judgeResult)||void 0===n?void 0:n.score)||0)-((null===(r=e.judgeResult)||void 0===r?void 0:r.score)||0)'", "/tmp/app.js chunk 65752", "HIGH", "HIGH — Sort order is fully confirmed. User rank override > AI score > judge score."),
    ("Grade mapping: C(≤4), B(5-6), A-(7-8), A(9-10)", "Grade function: 'x=e=>e<=4?{rank:\"C\"}:e<=6?{rank:\"B\"}:e<=8?{rank:\"A-\"}:{rank:\"A\"}'", "/tmp/app.js", "HIGH", "LOW — UI display grade, not backend ranking. Low impact."),
    ("Clips grouped by taskId (arcClipTaskId / clipCopilotTaskId)", "Clip grouping code: 'let s=e.arcClipTaskId||e.clipCopilotTaskId;s?t.has(s)?...:t.set(s,[u(e,1)])'", "/tmp/app.js chunk 65752", "HIGH", "HIGH — Multiple clip tasks run in parallel within a project. Each task produces its own ranked set."),
    ("pos_rate = rank/total per task", "pos_rate=Number((e.rank/s).toFixed(4)) where s = total clips in task group", "/tmp/app.js chunk 65752", "HIGH", "MEDIUM — Normalized position within task group."),
    ("isBonusClip flags extra clips", "isBonusClip:e.isBonusClip. Bonus clips excluded from pos_rate calculation.", "/tmp/app.js chunk 65752", "HIGH", "LOW — Bonus clip is additive."),
    ("Deduplication within task groups", "Clips grouped by taskId then pushed. Same taskId → same group → dedup within group.", "/tmp/app.js chunk 65752 sort + group logic", "HIGH", "MEDIUM — Grouping naturally deduplicates by task origin."),
    ("Rank can be manually overridden", "Sort function checks rank first. rank field overrides score.", "/tmp/app.js chunk 65752", "HIGH", "HIGH — User curation overrides AI when rank is set."),
    ("Diversification strategy unknown", "No evidence of explicit diversification logic. Possibly handled by task grouping.", "No evidence found", "LOW", "MEDIUM — Unknown if Opus explicitly diversifies or relies on grouping."),
    ("Reranking after edit unknown", "Clips can be edited (trim, extend, fix captions). Impact on ranking after edit unknown.", "No evidence found", "LOW", "LOW — Edit behavior likely affects final clip, not ranking."),
]

# ===================================================================
# RENDER_BACKEND_DEEP_DIVE
# ===================================================================
RENDER_BACKEND = [
    ["Claim", "Evidence", "Source", "Confidence", "Impact on Rewrite Decision"],
    ("Editing script with TRACKS = NLE compositor", "editingScript?.tracks — editing script has tracks property. This IS the NLE composition model.", "/tmp/app.js exportable-clips API", "HIGH", "CRITICAL — Editing script with tracks CONFIRMS track-based compositing. Not ffmpeg filter graph."),
    ("Frontend uses WASM (AVEditorEngine)", "AVEditorEngine-20260604-09125d5a.wasm.gz (5.96MB) on CDN. Loaded for editor and template pages.", "CDN + /tmp/dashboard.js", "HIGH", "HIGH — Frontend preview uses WASM-based compositor."),
    ("Zero ffmpeg in frontend", "0 hits in all frontend bundles (2.7MB JS + WASM)", "/tmp/app.js, dashboard.js, common-ui.js, main.js", "HIGH", "HIGH — Frontend compositing is 100% WASM, not ffmpeg."),
    ("Server-side render pipeline: Engine API", "5 engine-clips API endpoints: PUT/{id}, POST/render, POST/save-and-render, POST/refine, GET/{id}", "/tmp/dashboard.js", "HIGH", "HIGH — Server-side async rendering pipeline with polling."),
    ("Output: 4 formats including Adobe XML", "renderAsVideoPreview, renderAsVideoFile, renderAsVideoFile4K, renderAsAdobeXml", "/tmp/app.js clip data model", "HIGH", "MEDIUM — Multi-format output confirms professional workflow support."),
    ("Clip export tracking: clip-export-records", "POST/GET/PUT clip-export-records. POST check-export-operation. First-time export tracking.", "/tmp/app.js", "HIGH", "MEDIUM — Export monitoring and analytics."),
    ("Render elements identified", "'original-video-render', 'auto-tracking-video-render', 'auto-tracking-segment-render', 'add-section.mp4'", "/tmp/app.js CSS class names", "MEDIUM", "MEDIUM — Different render paths for different clip types."),
    ("Queue: fast_queue and no_credit_required", "'fast_queue' and 'no_credit_required' strings suggest priority queue tiers.", "/tmp/app.js", "MEDIUM", "MEDIUM — Queue priorities for different user tiers."),
    ("eq.opus.pro subdomain (Engine Queue?)", "eq.opus.pro and stg-eq.opus.pro subdomains discovered. 'eq' likely 'engine queue'.", "Bundle URL scan", "MEDIUM", "MEDIUM — Separate queue service for render jobs."),
    ("Server-side encoding technology UNKNOWN", "No evidence of server-side encoding method. Frontend = WASM. Backend could be ffmpeg or custom.", "No evidence", "LOW", "HIGH — Critical unknown for architecture decision."),
    ("B-roll generation with AI", "POST /brolls with sources: Stock (Pixels/Storyblocks/Shutterstock) or GenAi (AI generated). Approaches: Auto/Select/Prompt.", "/tmp/app.js brolls API", "HIGH", "MEDIUM — Opus has AI b-roll generation capability."),
]

# ===================================================================
# FEEDBACK_LOOP_DEEP_DIVE
# ===================================================================
FEEDBACK_LOOP = [
    ["Claim", "Evidence", "Source", "Confidence", "Impact on Rewrite Decision"],
    ("User feedback endpoint: like/dislike", "PUT /curated-clips/{clipId}.{archId}/feedback with body. Like/dislike tracked per clip.", "/tmp/app.js literal + endpoint", "HIGH", "HIGH — Users submit explicit feedback on clips. This data likely feeds back into scoring model."),
    ("Clip edit tracking (isEdited)", "isEdited = runId.endsWith('.CR'). Clips track whether they've been edited after generation.", "/tmp/app.js", "HIGH", "MEDIUM — User edits are tracked, potentially used as quality signal."),
    ("ArcQualityFeedback feature", "'ArcQualityFeedback' feature flag in feature list. User quality feedback collection.", "/tmp/app.js feature flag list", "MEDIUM", "MEDIUM — Quality feedback is a tracked feature, may influence model."),
    ("Like/dislike in clip list API", "like: e.like, dislike: e.dislike in clip data. f.dislike, a.f.isChecked tracked.", "/tmp/app.js clip data + like-clips endpoint", "HIGH", "MEDIUM — Explicit feedback stored with clip metadata."),
    ("GlobalNps feature", "'GlobalNps' in feature flag list. Net Promoter Score collection at app level.", "/tmp/app.js", "MEDIUM", "LOW — NPS is product feedback, not clip-specific feedback."),
    ("Engagement/retraining loop UNKNOWN", "No evidence of automatic retraining from engagement data. Likely exists but unconfirmed.", "No evidence found", "LOW", "HIGH — Whether Opus retrains from user feedback is critical unknown."),
    ("User preference learning UNKNOWN", "No evidence of personalized ranking from user behavior.", "No evidence found", "LOW", "MEDIUM — Personalized ranking could be a differentiator."),
    ("Feedback granularity: clip-level", "Feedback endpoint is per clip (clipId.archId). Feedback is at individual clip level, not project level.", "/tmp/app.js endpoint URL pattern", "HIGH", "LOW — Clip-level feedback allows fine-grained training data."),
]

# ===================================================================
# REWRITE_DECISION_READINESS
# ===================================================================
REWRITE_READINESS = [
    ["Unknown", "Current Confidence %", "Missing Evidence", "Can Decide Rewrite Yet?", "What We Need"],
    ("Candidate Generation Engine", "60%", "Exact chunking strategy unknown. Scene detection vs sliding window vs speaker-turn primary method unknown.", "NO", "Need to determine primary candidate boundary detection method."),
    ("Scoring Engine", "65%", "3 dims confirmed but weights unknown. LLM vs ML model unknown. Trend dim unconfirmed.", "NO", "Need to determine: (a) LLM or ML model, (b) weight distribution, (c) Trend dimension status."),
    ("Ranking Engine", "80%", "Sort order confirmed. Diversification unknown. Reranking after edit unknown.", "MOSTLY YES", "Ranking engine is 80% understood. Remaining unknowns are minor."),
    ("Render Backend", "55%", "Editing script with tracks confirms NLE compositing. Server-side encoding method unknown. Custom vs ffmpeg unknown.", "NO", "Need to determine: server-side encoding method. ffmpeg vs custom vs hybrid."),
    ("Feedback Loop", "35%", "like/dislike endpoint confirmed. Retraining loop unconfirmed. Personalization unknown.", "NO", "Need to determine: (a) retraining exists, (b) frequency, (c) personalization."),
    
    ("OVERALL READINESS", "55%", "Can distinguish frontend architecture (NLE tracks + WASM) = CRITICAL evidence. Cannot distinguish backend encoding method = CRITICAL gap.", "NO — NEED MORE EVIDENCE", "Top priority: Server-side encoding + rendering method. This determines whether GANYIQ architecture is fundamentally wrong or just needs refactoring."),
]

# Write ALL sheets
write_sheet("CANDIDATE_GENERATION_DEEP_DIVE", CANDIDATE_GENERATION[0], CANDIDATE_GENERATION[1:])
write_sheet("SCORING_ENGINE_DEEP_DIVE", SCORING_ENGINE[0], SCORING_ENGINE[1:])
write_sheet("RANKING_ENGINE_DEEP_DIVE", RANKING_ENGINE[0], RANKING_ENGINE[1:])
write_sheet("RENDER_BACKEND_DEEP_DIVE", RENDER_BACKEND[0], RENDER_BACKEND[1:])
write_sheet("FEEDBACK_LOOP_DEEP_DIVE", FEEDBACK_LOOP[0], FEEDBACK_LOOP[1:])
write_sheet("REWRITE_DECISION_READINESS", REWRITE_READINESS[0], REWRITE_READINESS[1:])

wb.save(PATH)
print(f"\n✅ Unknown elimination saved to: {PATH}")
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"\nRewrite decision readiness:")
for r in REWRITE_READINESS[1:]:
    print(f"  {r[0]:35s} | {r[1]:20s} | {r[3]:25s}")
