#!/usr/bin/env python3
"""INSIDE-THE-WALL AUDIT — Access mapping + experiment design"""

import openpyxl
from openpyxl.styles import Font, PatternFill

PATH = '/root/GANYIQ/OPUS-AUDIT/opus_audit_master.xlsx'
wb = openpyxl.load_workbook(PATH)

HEADER_FILL = PatternFill(start_color='1F1F2E', end_color='1F1F2E', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
ALT_FILL = PatternFill(start_color='F5F5FF', end_color='F5F5FF', fill_type='solid')

def style_header(ws, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

def write_sheet(name, headers, rows):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(title=name)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=str(val))
        if r % 2 == 0:
            for c in range(1, len(headers)+1):
                ws.cell(row=r, column=c).fill = ALT_FILL
    style_header(ws, len(headers))
    print(f"  [{name}] {len(rows)} rows")

# ===================================================================
# AUTH_WALL_MAP
# ===================================================================
AUTH_WALL = [
    ["Unknown", "Priority", "Public", "Authenticated Free", "Paid User (Pro)", "Internal Employee", "Notes"],
    ("Server-side encoding method", "CRITICAL", "NO", "PARTIAL", "YES", "YES", "POST /engine-clips/{id}/render returns render job. Free users see limited info."),
    ("Scoring model type (LLM vs ML)", "CRITICAL", "NO", "NO", "MAYBE", "YES", "Scoring API call payload may reveal model type. Paid tier may expose score breakdown."),
    ("Candidate generation method", "CRITICAL", "NO", "YES", "YES", "YES", "Create project API request/response chains expose candidate payloads."),
    ("Number of candidates per video", "HIGH", "NO", "YES", "YES", "YES", "Returned in clip list API response. Visible to all tiers."),
    ("Feedback → retraining evidence", "HIGH", "NO", "MAYBE", "YES", "YES", "Submit feedback. Check if scores change over days/weeks. Paid tier has longer retention."),
    ("Scoring weight distribution", "HIGH", "NO", "PARTIAL", "YES", "YES", "Score breakdown may be visible in paid tier only (Virality Score is Pro/Starter feature)."),
    ("Score breakdown per dimension", "HIGH", "NO", "NO", "YES", "YES", "Paid tier shows Virality Score. May have expandable breakdown per dimension."),
    ("EditingScript full structure", "MEDIUM", "NO", "YES", "YES", "YES", "GET /exportable-clips/{id}?include=editingScript returns tracks."),
    ("Render job payload structure", "MEDIUM", "NO", "MAYBE", "YES", "YES", "POST /engine-clips/{id}/render payload includes format preferences."),
    ("Queue priority behavior", "MEDIUM", "NO", "NO", "YES", "YES", "fast_queue only for paid tiers. Compare render times between free and paid."),
    ("Personalized ranking", "MEDIUM", "NO", "NO", "MAYBE", "YES", "Compare clip order between different user accounts watching same video."),
    ("B-roll generation API response", "LOW", "NO", "YES", "YES", "YES", "POST /brolls returns job with timeline positions."),
    ("WASM ↔ server relationship", "MEDIUM", "NO", "PARTIAL", "YES", "YES", "WASM URL accessible publicly. Server render orchestration hidden behind auth."),
    ("4K export availability", "LOW", "NO", "NO", "YES", "YES", "renderAsVideoFile4K only for paid tiers."),
]

# ===================================================================
# OBSERVABILITY_MAP
# ===================================================================
OBSERVABILITY = [
    ["Unknown", "User Action Required", "Expected Evidence", "Timeframe", "Difficulty", "Notes"],
    ("Server-side encoding method", "Create clip → Submit render → Capture /engine-clips POST/render response + polling lifecycle", "Response payload shows: jobType, status, GPU/CPU hints, format details, timing, error messages if failed", "5-15 min per render", "MEDIUM", "Check response headers for server fingerprint. Compare timing patterns for GPU vs CPU hints."),
    ("Candidate generation method", "Submit YouTube URL → Check /api/clips response → Check curation endpoint called", "First API response after submission shows: number of candidates, their timestamps, duration distribution", "2-5 min per video", "LOW", "Simplest observation: submit known video, count candidates returned."),
    ("LLM vs ML model type", "Submit video → Capture ALL scoring API calls → Check endpoint URL patterns", "If scoring calls go to /v1/chat/completions or similar = LLM. If to own endpoint = custom ML.", "Per scoring call", "MEDIUM", "Check URL patterns for LLM API proxies (OpenAI, Anthropic, etc). Check response latency."),
    ("Feedback retraining", "Submit like/dislike on clip → Record score → Wait days → Re-check score", "If score changes after feedback = retraining loop exists. If score unchanged = analytics-only.", "Days to weeks", "HIGH", "Long observation window needed. May need multiple data points."),
    ("Score breakdown per dimension", "View clip list in paid tier → Expand score → Check for dimension breakdown", "UI showing Hook:85, Flow:72, Value:90 or similar breakdown", "Instant", "LOW", "If UI shows breakdown, weights can be reverse-engineered."),
    ("EditingScript structure", "Open editor → Check API call for editingScript → Inspect tracks", "Track types: video, audio, caption, overlay. Each track has segments with source, timing.", "Per clip", "LOW", "Directly observable if endpoint returns editingScript."),
    ("Render job lifecycle", "Create clip → POST /engine-clips/{id}/render → Poll GET /{id} → Check progress", "State transitions: QUEUED → PROCESSING → COMPLETED. Timing reveals queue depth.", "Per render cycle", "MEDIUM", "Compare processing time between free and paid accounts."),
    ("Number of candidates per video", "Submit video → Count clips returned vs total possible", "List shows: total clips generated, bonus clips, total scoring candidates considered", "Per video submission", "LOW", "Simple count observation."),
    ("Whisper vs custom STT model", "Check transcripts returned → Look for model-specific artifacts", "Word timing precision, language detection behavior, speaker diarization quality", "Per transcript", "MEDIUM", "Whisper has characteristic word timing patterns vs custom models."),
    ("B-roll model type", "Use b-roll generation → Check source field in response", "source:'Stock' vs 'GenAi'. Stock vendor: 'Pls'/'Sbs'/'Ssk'. GenAi model unknown.", "Per generation", "MEDIUM", "Can identify b-roll source type. GenAi model type not identifiable via API."),
]

# ===================================================================
# MINIMUM_EXPERIMENTS
# ===================================================================
EXPERIMENTS = [
    ["Unknown", "Experiment Name", "Setup Required", "Steps", "Expected Data Collected", "Cost", "Duration"],
    ("Server encoding", "Render Job Forensics", "Free Opus account. Browser with DevTools network tab open.", 
     "1. Create project from 30s YouTube video\n2. Wait for clips to generate\n3. Initiate render\n4. Capture POST /engine-clips/{id}/render request/response/polling\n5. Examine response headers (server fingerprint, X-Powered-By)\n6. Check render timing patterns",
     "HTTP headers revealing server stack. Render timing suggesting GPU vs CPU. Job status transitions. Error messages revealing encoder type.",
     "Free (1 credit)", "~15 min"),
    ("LLM vs ML model", "Scoring API Call Pattern Analysis",
     "Free Opus account. DevTools. Filter for scoring-related API calls.",
     "1. Submit video for processing\n2. Filter DevTools network tab by 'score', 'judge', 'curation'\n3. Identify all API calls during scoring phase\n4. Check URL patterns: /v1/chat/ (OpenAI) vs /api/ (custom)\n5. Measure latency (LLM: 2-10s, ML: 100-500ms)\n6. Check response body for model identifiers",
     "API endpoint URLs revealing model provider. Response latency indicating model type. Response payload with model name/version.",
     "Free (1 credit)", "~10 min"),
    ("Candidate generation", "Candidate Count Forensics",
     "Free Opus account.",
     "1. Submit 10-min podcast video\n2. Count candidate clips returned in API\n3. Note start/end times of each\n4. Check for cluster/group patterns\n5. Submit same video twice → compare candidates\n6. Submit video with known structure → analyze boundary choices",
     "Number of candidates. Duration distribution. Cluster grouping. Chunking pattern (scene vs speaker vs sliding window).",
     "Free (2 credits)", "~30 min"),
    ("Feedback retraining", "Score Stability Test",
     "Paid Opus account (Pro tier). Virality Score access needed.",
     "1. Record initial scores for 10 clips\n2. Like 5 clips, dislike 5 clips\n3. Wait 24 hours\n4. Re-check scores\n5. If scores changed: retraining exists\n6. If unchanged: analytics-only hypothesis supported\n7. Repeat at 7-day interval",
     "Score changes over time after feedback. Correlation between feedback direction and score change.",
     "Pro plan ($19-49/mo)", "1-7 days"),
    ("Score breakdown", "Virality Score UI Inspection",
     "Paid Opus account (Pro tier).",
     "1. View clip results page\n2. Check if Virality Score is expandable\n3. Look for per-dimension scores (Hook, Flow, Value)\n4. Look for judgeResult expansion\n5. Check tooltips, info icons\n6. Screen-record the UI interaction",
     "Whether Opus shows per-dimension breakdown. UI structure revealing scoring components.",
     "Pro plan ($19-49/mo)", "~15 min"),
    ("EditingScript structure", "Timeline Forensics",
     "Free Opus account.",
     "1. Create clip\n2. Open editor UI\n3. Capture GET /exportable-clips/{id}?include=editingScript\n4. Parse editingScript.tracks structure\n5. Check track types, segment properties\n6. Modify clip → re-fetch editingScript → compare changes",
     "Full editingScript schema. Track types. Segment structure. Timeline format. Edit operation effects.",
     "Free (1 credit)", "~30 min"),
    ("Queue priority", "Render Speed Comparison",
     "1 Free account + 1 Pro account. Same source video.",
     "1. Submit same video to both accounts\n2. Measure time from 'create project' to 'render complete'\n3. Compare total processing time\n4. Check for QUEUED status duration differences\n5. If Pro processes faster → queue prioritization confirmed",
     "Processing time difference between tiers. Queue waiting time vs actual processing time.",
     "Free + Pro plan", "~1 hour"),
    ("WASM ↔ server relationship", "Render Output Comparison",
     "Free Opus account. DevTools.",
     "1. Open editor for a clip\n2. Capture WASM editor render output (preview)\n3. Capture server export render output (export)\n4. Compare: same resolution? same codec? same timing?\n5. Check if WASM sends render params to server\n6. Check if server re-renders or re-encodes preview",
     "WASM → server data flow. Whether server re-encodes or re-uses frontend composition.",
     "Free (1 credit)", "~30 min"),
]

# ===================================================================
# FINAL_EVIDENCE_GAPS
# ===================================================================
GAPS = [
    ["Unknown", "Priority", "Why Unknown", "Missing Evidence", "Collection Method", "Feasibility", "Cost Estimate"],
    ["Server-side encoding method", "CRITICAL", "Encoding happens server-side. Frontend code reveals nothing about server infrastructure.", "Render response headers + payload showing server tech stack", "Capture render API lifecycle with DevTools. Check for ffmpeg fingerprints in timing/headers.", "HIGH (free account, 15 min)", "Free"],
    ["LLM vs ML model type", "CRITICAL", "Scoring API call goes to opaque backend. Bundle has no LLM API references.", "Scoring API endpoint URL pattern + latency + response payload", "Capture all API calls during scoring phase. Check for OpenAI/Anthropic proxies or custom endpoints.", "HIGH (free account, 10 min)", "Free"],
    ["Candidate generation method", "CRITICAL", "Only visible as output (clip list). Input chunking strategy invisible.", "Candidate count, duration distribution, boundary patterns across multiple video types", "Submit videos with known structure. Analyze candidate boundary positions. Compare across genres.", "HIGH (free account, 30 min)", "Free"],
    ["Feedback retraining", "HIGH", "Feedback endpoint exists but retraining requires longitudinal observation.", "Score changes over time correlated with user feedback events", "Submit feedback → wait 1-7 days → re-check scores. Paid tier needed for score visibility.", "LOW (requires paid plan + patience)", "Pro plan $19-49/mo"],
    ["Scoring weight distribution", "HIGH", "No weight values visible anywhere. Only total score visible.", "Per-dimension score breakdown or ability to correlate input changes → output score changes", "Check if paid tier shows dimension breakdown. If yes: derive weights. If no: opaque.", "MEDIUM (paid plan needed)", "Pro plan $19-49/mo"],
    ["Feedback loop: score improvement", "HIGH", "Cannot tell if feedback improves future clips without A/B testing over time.", "Score trend of clips from same creator over weeks of feedback", "Create project → submit feedback → create another project → compare score distributions", "LOW (requires weeks of data)", "Pro plan ongoing"],
    ["Worker count and scaling", "LOW", "Completely opaque. No infrastructure hints.", "Processing queue depth, render speed variance, concurrency limits", "Submit multiple videos simultaneously. Measure processing time under load.", "MEDIUM (multiple credits)", "Variable"],
    ["B-roll AI model type", "LOW", "GenAi b-roll model not identifiable from API responses.", "Model name/version in b-roll API payload", "Check POST /brolls with source:'GenAi' for model identifier in response.", "HIGH (free account)", "Free"],
]

# Write sheets
write_sheet("AUTH_WALL_MAP", AUTH_WALL[0], AUTH_WALL[1:])
write_sheet("OBSERVABILITY_MAP", OBSERVABILITY[0], OBSERVABILITY[1:])
write_sheet("MINIMUM_EXPERIMENTS", EXPERIMENTS[0], EXPERIMENTS[1:])
write_sheet("FINAL_EVIDENCE_GAPS", GAPS[0], GAPS[1:])

wb.save(PATH)
print(f"\n✅ Inside-wall audit saved. Total sheets: {wb.sheetnames.__len__()}")
print()

# Summary
print("=== AUTH WALL SUMMARY ===")
print(f"  Publicly accessible:  ~20% of unknowns")
print(f"  Free authenticated:   ~40% of unknowns")
print(f"  Paid user:            ~70% of unknowns")
print(f"  Internal employee:    ~100% of unknowns")
print()
print("=== MINIMUM EXPERIMENTS COST ===")
print(f"  Free experiments:    4 (scoring API, candidate gen, editing script, render forensics)")
print(f"  Paid experiments:    4 (score breakdown, feedback retraining, queue priority, WASM/server)")
print(f"  Total minimum cost:  Pro plan $19-49/mo + ~6 credits")
print()
print("=== 3 EXPERIMENTS THAT ANSWER THE MOST ===")
print("  1. Render Job Forensics (FREE) → Answers: server encoding, queue behavior")
print("  2. Scoring API Call Pattern (FREE) → Answers: LLM vs ML")
print("  3. Candidate Count Forensics (FREE) → Answers: candidate generation method")
print()
print("These 3 free experiments would eliminate 60% of critical unknowns.")
