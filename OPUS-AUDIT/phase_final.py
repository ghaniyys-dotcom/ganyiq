#!/usr/bin/env python3
"""PHASE FINAL — Unknown elimination with evidence-only findings"""

import openpyxl
from openpyxl.styles import Font, PatternFill

PATH = '/root/GANYIQ/OPUS-AUDIT/opus_audit_master.xlsx'
wb = openpyxl.load_workbook(PATH)

HEADER_FILL = PatternFill(start_color='1F1F2E', end_color='1F1F2E', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
ALT_FILL = PatternFill(start_color='F5F5FF', end_color='F5F5FF', fill_type='solid')

def style_header(ws, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

def write_sheet(name, headers, rows):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(title=name)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=str(val))
        if r % 2 == 0:
            for c in range(1, len(headers)+1):
                ws.cell(row=r, column=c).fill = ALT_FILL
    style_header(ws, len(headers))
    ws.auto_filter.ref = f"A1:{chr(64+len(headers))}{len(rows)+1}"
    print(f"  {name}: {len(rows)} rows")

# ===================================================================
# CANDIDATE_GENERATION_FORENSICS_V2
# ===================================================================
CANDIDATE_V2 = [
    ["Claim", "Evidence", "Source", "Classification", "Confidence", "Notes"],
    ("Cluster-based candidate organization", "clips have 'cluster' and 'clusterJobId' fields. POST /cluster-jobs with jobType:'GEN_TITLE'. GET /cluster-jobs/{id}/progress?ms= tracks progress.", "/tmp/app.js literal: cluster, clusterJobId, cluster-jobs endpoint", "FACT", "HIGH", "Clips are organized into clusters. Cluster jobs generate titles per cluster."),
    ("Screenplay with chapters", "screenplay field in clip data model has chapters array. Each chapter has lines with words.", "/tmp/app.js s() function: screenplay.e.chapters", "FACT", "HIGH", "Screenplay is structured as chapters. Each chapter contains transcript lines with word-level data."),
    ("transcriptSegments in API calls", "transcriptSegments field passed to voice-over API alongside screenplay", "/tmp/app.js voice-over API call", "FACT", "HIGH", "Transcript is pre-segmented before API processing."),
    ("User-defined clip durations", "clipDurations array in curationPref. Users control desired clip lengths.", "/tmp/app.js + help.opus.pro docs", "FACT", "HIGH", "Duration influences how candidates are filtered."),
    ("topicKeywords filter candidates", "topicKeywords in curationPref. Official docs say keywords must be in transcript.", "Bundle + official docs", "FACT", "HIGH", "Keywords filter candidates by transcript match."),
    ("Genre-specific curation models", "curation_api_config with genres ready[] and beta[]. Genre field in clip data.", "/tmp/app.js", "FACT", "HIGH", "Different models per genre."),
    ("Visual timeline editor", "User can 'Arrange Scenes in Timeline', drag segment boundaries. 'add-section-screenplay' CSS class.", "/tmp/app.js string literals", "FACT", "HIGH", "Visual timeline editing."),
    ("Candidate generation PRIMARY METHOD UNKNOWN", "No evidence of sliding window, speaker-turn boundaries, scene detection, or any specific chunking algorithm.", "N/A", "SPECULATION", "LOW", "Still don't know HOW candidates are initially generated from raw video."),
    ("Transcript → Chapter → Cluster → Clip hypothesis", "Combined evidence: transcriptSegments → screenplay.chapters → clusterJobs → individual clips", "Multiple sources", "INFERENCE", "MEDIUM", "Likely pipeline: video → transcript → chapters → clusters → candidate clips → scored → ranked."),
]

# ===================================================================
# FEEDBACK_LOOP_FORENSICS_V2
# ===================================================================
FEEDBACK_V2 = [
    ["Claim", "Evidence", "Source", "Classification", "Confidence", "Notes"],
    ("Feedback endpoint EXISTS", "PUT /curated-clips/{clipId}.{archId}/feedback with body", "/tmp/app.js literal", "FACT", "HIGH", "Endpoint confirmed. Feedback can be submitted."),
    ("Like/dislike tracked per clip", "like:e.like, dislike:e.dislike in clip data. f.dislike tracked.", "/tmp/app.js like-clips endpoint", "FACT", "HIGH", "Explicit feedback (like/dislike) stored per clip."),
    ("isEdited clip tracking", "isEdited = runId.endsWith('.CR'). Clips track edit state.", "/tmp/app.js", "FACT", "HIGH", "User edits are tracked."),
    ("ArcQualityFeedback feature flag", "ArcQualityFeedback in feature flag list. Quality feedback feature.", "/tmp/app.js Statsig config", "FACT", "MEDIUM", "Quality feedback is a managed feature."),
    ("ZERO evidence of retraining from feedback", "retrain: 0 hits across all bundles. No ML training references.", "All bundles", "FACT (absence)", "HIGH", "No evidence whatsoever that feedback is used for model retraining."),
    ("ZERO evidence of engagement data in scoring", "engagement (15 hits) — all marketing text. No scoring engine references.", "/tmp/app.js", "FACT (absence)", "HIGH", "Engagement is discussed in marketing context only."),
    ("ZERO evidence of personalized ranking", "No user preference → ranking personalization evidence found.", "All bundles", "FACT (absence)", "HIGH", "No personalized clip ranking."),
    ("Feedback appears analytics-only", "No counter-evidence found that feedback affects models or rankings.", "All bundles", "INFERENCE", "MEDIUM", "Feedback endpoints exist but appear to be for UI display and analytics, not ML training."),
]

# ===================================================================
# BACKEND_RENDER_FORENSICS_V2
# ===================================================================
BACKEND_V2 = [
    ["Claim", "Evidence", "Source", "Classification", "Confidence", "Notes"],
    ("ffmpeg in frontend: ZERO", "0 hits ffmpeg across app.js, dashboard.js, common-ui.js, main.js, chunk9201.js", "All 5 bundles", "FACT", "HIGH", "Frontend has zero ffmpeg references."),
    ("WASM engine exists on CDN", "AVEditorEngine-20260604-09125d5a.wasm.gz (5.96MB) at public.cdn.opus.pro/clip-web/wasm/", "CDN probe", "FACT", "HIGH", "Frontend compositor is WASM-based."),
    ("Editing script with tracks", "editingScript?.tracks — NLE-style track-based composition format", "/tmp/app.js exportable-clips API", "FACT", "HIGH", "Track-based editing script format."),
    ("GPU in frontend: CSS only", "gpu (2 hits): 'transform-gpu' CSS class. Zero GPU computing references.", "/tmp/app.js", "FACT", "HIGH", "GPU references are CSS rendering, not video processing."),
    ("NVENC: ZERO evidence", "nvenc: 0 hits across all bundles", "All bundles", "FACT (absence)", "HIGH", "No NVENC references anywhere."),
    ("CUDA: ZERO evidence", "cuda: 0 hits across all bundles", "All bundles", "FACT (absence)", "HIGH", "No CUDA references anywhere."),
    ("Encoder: CSS/base64 only", "encoder (3 hits): all in CSS transform or base64 encoding context", "/tmp/app.js", "FACT", "HIGH", "No video encoder references."),
    ("Server encoding method: COMPLETELY UNKNOWN", "No evidence of ffmpeg, WASM, custom encoder, or any encoding method on server-side", "All sources", "SPECULATION", "LOW", "Most critical unknown. Cannot determine server rendering architecture from frontend evidence."),
    ("Engine API endpoints exist (5)", "PUT/{id}, POST/render, POST/save-and-render, POST/refine, GET/{id} for engine-clips", "/tmp/dashboard.js", "FACT", "HIGH", "Server-side render API with async pattern."),
    ("4 export formats", "renderAsVideoPreview, VideoFile, VideoFile4K, AdobeXml + URI variants", "/tmp/app.js", "FACT", "HIGH", "Multi-format output."),
    ("Queue priority tiers", "'fast_queue' and 'no_credit_required' strings. eq.opus.pro (Engine Queue?) subdomain.", "/tmp/app.js + subdomain probe", "FACT", "MEDIUM", "Priority queue for different processing tiers."),
]

# ===================================================================
# SCORING_ENGINE_VALIDATION_V2
# ===================================================================
SCORING_V2 = [
    ["Claim", "Evidence", "Source", "Classification", "Confidence", "Notes"],
    ("3 dimensions confirmed in official docs", "Hook: 'grab attention'. Flow: 'logical flow'. Value: 'resonate emotionally and create personal connection'.", "help.opus.pro virality-score page", "FACT", "HIGH", "All 3 described in official documentation. Value = emotional connection, NOT actionable advice."),
    ("Score range 0-99", "Official docs: 'Virality score ranges from 0 to 99'", "help.opus.pro virality-score page", "FACT", "HIGH", "Directly stated."),
    ("Default sort by score descending", "Official docs: 'By default, your clips will be sorted by Virality Score, from highest to lowest.'", "help.opus.pro virality-score page", "FACT", "HIGH", "Sort behavior confirmed."),
    ("judgeResult object exists", "judgeResult:{isCopilotClip,relevanceScore,trendTopic,score} in clip data model", "/tmp/app.js", "FACT", "HIGH", "Contains additional scoring metadata."),
    ("Sort function: rank > score > judgeResult.score", "Sort function literal in bundle. rank first, score second, judgeResult.score tiebreaker.", "/tmp/app.js chunk 65752", "FACT", "HIGH", "Sort order fully confirmed."),
    ("4 score keys in bundle (hookScore, coherenceScore, connectionScore, trendScore)", "Literal strings: hookScore, coherenceScore, connectionScore, trendScore with commentKey fields", "/tmp/app.js UI config", "FACT", "HIGH", "4 score keys exist in UI but only 3 in official docs."),
    ("Trend dimension is NOT in official docs", "Official docs list ONLY Hook+Flow+Value. Trend not mentioned.", "help.opus.pro virality-score page", "FACT", "HIGH", "4th dimension (Trend) is bundle-only. Not confirmed as official scoring dimension."),
    ("LLM model type: UNCONFIRMED", "Docs say 'AI evaluates'. Bundle has 0 LLM API call references.", "All sources", "SPECULATION", "LOW", "Cannot determine if scoring uses LLM or smaller ML model."),
    ("Scoring weights: UNKNOWN", "No evidence of weight values for Hook, Flow, Value, or Trend.", "All sources", "UNKNOWN", "N/A", "No evidence. Remove all previous weight estimates."),
    ("Prompt relevance modifier CONFIRMED", "Official docs: 'evaluates whether the clip is relevant to your prompt, when using ClipAnything model'", "help.opus.pro virality-score page", "FACT", "HIGH", "Additional modifier for ClipAnything mode."),
    ("Score is server-side generated", "No scoring logic in frontend. Scores provided by API.", "All bundles", "FACT", "HIGH", "Scoring is entirely server-side."),
    ("Premium feature", "Official docs: 'available exclusively to Pro and Starter plan users'", "help.opus.pro virality-score page", "FACT", "HIGH", "Free users don't see scores."),
]

# ===================================================================
# TOP_25_UNKNOWNS_REMAINING
# ===================================================================
TOP_25_UNKNOWNS = [
    ["Rank", "Unknown", "Importance", "Architectural Impact", "Evidence Gap", "Can Resolve With"],
    [1, "Server-side encoding method (ffmpeg/custom/hybrid)", "CRITICAL", "DETERMINES REWRITE vs RESTRUCTURE", "ZERO evidence. All frontend. No ffmpeg, no encoder, no GPU references.", "Capture /engine-clips render API response. Check headers, job payloads."],
    [2, "Primary candidate generation method", "CRITICAL", "HIGH — determines if GANYIQ chunking is correct", "Transcript segments, chapters, clusters exist but primary method unknown.", "Create Opus account. Submit video. Observe API calls for candidate generation payload."],
    [3, "ML model type (LLM vs BERT vs CNN)", "CRITICAL", "HIGH — determines AI infrastructure cost/feasibility", "Docs say 'AI' not 'LLM'. No model identification possible.", "Capture scoring API calls. Check endpoint patterns for LLM API proxies."],
    [4, "Feedback → model retraining loop", "HIGH", "HIGH — determines data moat depth", "Feedback endpoint EXISTS but ZERO evidence of retraining. Analytics-only likely.", "Submit feedback, check if scores change. A/B test over time."],
    [5, "Scoring weight distribution (Hook/Flow/Value weights)", "HIGH", "MEDIUM — affects GANYIQ scoring accuracy", "ZERO evidence for any weight values. All previous estimates removed.", "Compare scores across known clip types. Reverse-engineer via behavior analysis."],
    [6, "Number of candidates generated per video", "MEDIUM", "MEDIUM — affects processing capacity planning", "No evidence of candidate count limits.", "Create account. Submit test video. Count clips returned."],
    [7, "Server-side database technology", "LOW", "LOW — doesn't affect architecture decision", "No evidence.", "Not needed for architecture decision."],
    [8, "Server-side queue technology", "LOW", "LOW — doesn't affect architecture decision", "No evidence.", "Not needed for architecture decision."],
    [9, "GPU/NVENC on server-side", "MEDIUM", "MEDIUM — affects encoding cost model", "ZERO evidence in frontend. Server-side unknown.", "Check render job timing for GPU vs CPU indicators."],
    [10, "Deduplication strategy", "MEDIUM", "LOW — minor optimization", "Grouping by taskId inferred. Exact strategy unknown.", "Not needed for initial architecture decision."],
    [11, "Diversification strategy", "LOW", "LOW — nice to have", "No evidence.", "Not needed for architecture decision."],
    [12, "Personalization/ranking per user", "MEDIUM", "MEDIUM — product decision", "No evidence of personalized ranking.", "Check if scores/rank change between user sessions."],
    [13, "Human review process", "MEDIUM", "MEDIUM — quality assurance", "No evidence.", "Check for manual approval workflows."],
    [14, "Model update/training frequency", "MEDIUM", "MEDIUM — affects competitive timeline", "No evidence.", "Track score consistency over time."],
    [15, "Storage strategy (hot/cold tier)", "LOW", "LOW — infrastructure detail", "GCS confirmed. Tier unknown.", "Not needed for architecture decision."],
    [16, "ClipAnything retrieval mechanism", "MEDIUM", "MEDIUM — feature parity", "Doc describes what, not how.", "Check ClipAnything API payloads for embedding vs keyword approach."],
    [17, "Language model for emphasis/keyword detection", "MEDIUM", "LOW — subtitle feature", "emphasis-engine in GANYIQ. Opus equivalent unknown.", "Not critical for initial decision."],
    [18, "Hosting provider (AWS/GCP/Azure)", "LOW", "LOW — infrastructure detail", "GCS confirmed (GCP). Compute unknown.", "Not needed for architecture decision."],
    [19, "Worker count and scaling strategy", "LOW", "LOW — infrastructure detail", "No evidence.", "Not needed for architecture decision."],
    [20, "Pricing per clip cost structure", "LOW", "LOW — business decision", "API = 1 credit/minute known.", "Not needed for architecture decision."],
    [21, "Auto-layout ML model type", "MEDIUM", "LOW — 8 layouts already known", "Layouts known. Model type unknown.", "Not critical for initial architecture decision."],
    [22, "Brand template rendering approach", "LOW", "LOW — cosmetic", "Templates known. Rendering method unknown.", "Not needed."],
    [23, "B-roll generation model type", "MEDIUM", "LOW — feature gap", "Stock + GenAi known. Model type unknown.", "Not critical for initial decision."],
    [24, "B-roll integration into timeline", "MEDIUM", "LOW — feature gap", "selectedTimeline, editedTimeline fields known.", "Not critical for initial decision."],
    [25, "WASM → server render relationship", "HIGH", "MEDIUM — determines frontend/backend split", "WASM for preview. Server for export. Exact split unknown.", "Check if server uses same WASM compositor or different encoding."],
]

# Write ALL sheets
write_sheet("CANDIDATE_GENERATION_FORENSICS_V2", CANDIDATE_V2[0], CANDIDATE_V2[1:])
write_sheet("FEEDBACK_LOOP_FORENSICS_V2", FEEDBACK_V2[0], FEEDBACK_V2[1:])
write_sheet("BACKEND_RENDER_FORENSICS_V2", BACKEND_V2[0], BACKEND_V2[1:])
write_sheet("SCORING_ENGINE_VALIDATION_V2", SCORING_V2[0], SCORING_V2[1:])
write_sheet("TOP_25_UNKNOWNS_REMAINING", TOP_25_UNKNOWNS[0], TOP_25_UNKNOWNS[1:])

wb.save(PATH)
print(f"\n✅ Phase final saved to: {PATH}")
print(f"Total sheets: {len(wb.sheetnames)}")

print("\n=== ANSWERS TO 5 PRIMARY OBJECTIVES ===")
print()
print("Q1: How Opus finds candidate clips?")
print("  A: Cluster-based approach. Video → transcript → screenplay/chapters → cluster jobs → clips.")
print("  Classification: INFERENCE (MEDIUM). Primary chunking method still unknown.")
print()
print("Q2: How Opus scores candidate clips?")
print("  A: 3-dimension AI evaluation (Hook + Flow + Value). Trend is bundle-only.")
print("  Score 0-99. Sort: rank > score > judgeResult.score.")
print("  Classification: FACT (HIGH) for 3 dims. LLM vs ML = UNKNOWN.")
print()
print("Q3: How Opus renders clips?")
print("  A: Frontend: WASM-based NLE with editingScript?.tracks. Server: COMPLETELY UNKNOWN.")
print("  ZERO ffmpeg, ZERO GPU, ZERO NVENC, ZERO CUDA evidence anywhere.")
print("  Classification: Frontend = FACT (HIGH). Backend = UNKNOWN.")
print()
print("Q4: Does feedback improve the model?")
print("  A: NO EVIDENCE. Feedback endpoint EXISTS. ZERO evidence of retraining, personalization, or ML usage.")
print("  Classification: Feedback collection = FACT. Feedback usage = SPECULATION (likely analytics-only).")
print()
print("Q5: Does Opus use ffmpeg, custom rendering, or hybrid?")
print("  A: Frontend = ZERO ffmpeg (WASM NLE). Backend = COMPLETELY UNKNOWN.")
print("  Cannot determine server encoding method from frontend evidence alone.")
