#!/usr/bin/env python3
"""Update opus_audit_master.xlsx with ALL new forensic findings."""

import openpyxl
from openpyxl.styles import Font, PatternFill
from copy import copy

PATH = '/root/GANYIQ/OPUS-AUDIT/opus_audit_master.xlsx'
wb = openpyxl.load_workbook(PATH)

HEADER_FILL = PatternFill(start_color='1F1F2E', end_color='1F1F2E', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
ALT_FILL = PatternFill(start_color='F5F5FF', end_color='F5F5FF', fill_type='solid')

def style_header(ws, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

def write_sheet(name, headers, rows):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(title=name)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=str(val))
        if r % 2 == 0:
            for c in range(1, len(headers)+1):
                ws.cell(row=r, column=c).fill = ALT_FILL
    style_header(ws, len(headers))
    ws.auto_filter.ref = f"A1:{chr(64+len(headers))}{len(rows)+1}"
    for c in range(1, len(headers)+1):
        ws.column_dimensions[chr(64+c)].width = max(12, min(50, headers[c-1].__len__() * 1.5 + 5))
    print(f"  {name}: {len(rows)} rows")

# ===================================================================
# SHEET 1: NETWORK_FORENSICS
# ===================================================================
NETWORK_FORENSICS = [
    ["Endpoint", "Method", "Payload/Params", "Response", "Purpose", "Evidence Source", "Confidence", "Notes"],
    ["/api/clip-projects", "POST", '{"videoUrl","curationPref","importPref","brandTemplateId","layoutOpts","captionOpts"}', "Project created, returns projectId", "Create clipping project from YouTube URL", "Bundle: pages/dashboard JS + Mintlify docs", "HIGH", "Main entry point for all clip processing"],
    ["/api/clips", "GET", "?q=findByProjectId&projectId=X", "List of clips with scores, ranks, URIs", "List generated clips for a project", "Bundle: app JS + Mintlify docs", "HIGH", "Returns full clip data model with scores"],
    ["/api/transcripts", "GET", "?q=findByProjectId&projectId=X", "Transcript with paragraphs, speakers, word timing", "Fetch transcript with speaker labels", "Bundle: app JS", "HIGH", "Per-word timing available for caption sync"],
    ["/api/censor-jobs", "POST", "clipId, type", "Job created, returns jobId", "Auto-censor content (blur/faces)", "Mintlify docs", "MEDIUM", "Batch processing with status polling"],
    ["/api/censor-jobs/{jobId}", "GET", "-", "Job status + result", "Poll censor job", "Mintlify docs", "MEDIUM", "Async pattern"],
    ["/api/generative-jobs", "POST", "clipId, style", "Job created", "Generate AI thumbnails", "Mintlify docs", "MEDIUM", "Async thumbnail generation"],
    ["/api/generative-jobs/{jobId}", "GET", "-", "Job status + image URL", "Poll thumbnail job", "Mintlify docs", "MEDIUM", ""],
    ["/api/collections", "POST", "name, clipIds", "Collection created", "Group clips into collections", "Mintlify docs", "HIGH", "Batch organization feature"],
    ["/api/collections/{id}/export", "POST", "ids[]", "Export job created", "Bulk export clips", "Mintlify docs", "HIGH", "Batch export feature"],
    ["/api/social-accounts", "GET", "-", "List connected platforms", "List connected social media", "Bundle: app JS", "HIGH", "Multi-platform publishing"],
    ["/api/social-copy-jobs", "POST", "clipId, platform", "Job created", "Generate social media caption", "Bundle: app JS + Mintlify", "HIGH", "AI-generated post copy"],
    ["/api/post-tasks", "POST", "clipId, platform, time", "Publish queued", "Publish clip immediately", "Bundle: app JS", "HIGH", "Direct social publishing"],
    ["/api/publish-schedules", "POST", "clipId, platform, scheduleTime", "Schedule created", "Schedule clip for later", "Bundle: app JS", "HIGH", "Scheduled publishing"],
    ["/api/brand-templates", "GET", "-", "List of brand templates", "Get available brand templates", "Bundle: app JS", "HIGH", "Template system for captions/layout"],
    ["/api/share-project", "POST", "projectId, visibility", "Updated", "Share/update project visibility", "Mintlify docs", "MEDIUM", "Collaboration feature"],
    ["/engine-clips/{clipId}.{archId}", "PUT", "render preferences", "Updated", "Update clip render settings", "Bundle: dashboard JS", "HIGH", "Engine-level clip config"],
    ["/engine-clips/{clipId}.{archId}/render", "POST", "format, resolution", "Render job queued", "Trigger render for a clip", "Bundle: dashboard JS", "HIGH", "Triggers actual video rendering"],
    ["/engine-clips/{clipId}.{archId}/save-and-render", "POST", "all prefs", "Render job queued", "Save prefs + trigger render", "Bundle: dashboard JS", "HIGH", "Combined save+render action"],
    ["/engine-clips/{clipId}.{archId}/refine", "POST", "refinement params", "Refine job queued", "Refine existing clip", "Bundle: dashboard JS", "MEDIUM", "Re-run AI selection on same clip"],
    ["/api/curation-api-config", "GET", "-", "Config with genres list", "Get available curation config", "Bundle: app JS", "HIGH", "Contains genres: ready[], beta[]"],
    ["/api/create-project", "POST", "videoUrl, importPref, curationPref", "Project created", "Alias for clip-projects endpoint", "Mintlify docs", "MEDIUM", "Older endpoint name"],
    ["/api/workers/{workerId}/heartbeat", "POST", "version", "200 OK", "Worker heartbeat", "Existing GANYIQ knowledge", "HIGH", "Same pattern as GANYIQ worker"],
    ["health/api.ping", "GET", "-", "pong", "Health check", "Bundle: network traces", "MEDIUM", "Internal health endpoint"],
]

# ===================================================================
# SHEET 2: OUTPUT_FORENSICS (from marketing + observed behavior)
# ===================================================================
OUTPUT_FORENSICS = [
    ["Category", "Observed Pattern", "Typical Duration", "Speaker Count", "Hook Position", "Layout", "Zoom Events", "Camera Switches", "Reaction Cuts", "Subtitle Style", "Evidence Source", "Confidence"],
    ["Podcast (2-person)", "Split 50/50 or 60/40 hero+reaction, speaker-focused swaps", "30-90s", "2", "First 3s", "Split 50/50, Hero+Reaction, Single", "1-2 slow zooms on active speaker", "Every 8-15s on speaker change", "Cut to listener on laughter/gasp", "Karaoke highlight, bottom, single word", "Bundle: DecisionEngine code analysis", "HIGH"],
    ["Podcast (3+ person)", "Split 33/33/33 or 4-way grid on group discussions", "30-60s", "3-4", "First 3s", "Split 3-way, Split 4-way", "Minimal zoom, mostly static grid", "Every 10-20s", "When non-speaker reacts", "Smaller font, bottom, highlight speaker", "Bundle: layout system analysis", "MEDIUM"],
    ["Business/Interview", "Single face dominant, occasional split. Clean professional look", "30-120s", "1-2", "First 2-5s", "Single (Fill/Fit)", "Slow Ken Burns zoom on single face", "Rare (every 15-30s)", "Listener PiP for reactions", "Clean karaoke, minimal styling", "Marketing examples", "HIGH"],
    ["Self Improvement", "Single speaker full frame, motivational pacing", "30-60s", "1", "Hook in first 2s", "Single (Fill/Fit)", "Very slow Ken Burns (0.04 zoom/sec)", "Almost none", "N/A (1 speaker only)", "Bold karaoke, large font, highlighted key words", "Bundle: zoompan params analysis", "HIGH"],
    ["Comedy", "Fast cuts, frequent layout switches, reaction-heavy", "15-45s", "2+", "Immediate punchline", "Single → Reaction → Split rapidly", "Frequent punch-in on reaction", "Every 3-8s", "Many reaction cuts", "Emoji support, animated captions", "Bundle: ReactionScheduler analysis", "MEDIUM"],
    ["Gaming", "Game layout (unique): gameplay top + facecam bottom", "30-60s", "1-2", "Key moment", "Game layout (unique)", "None (fixed layout)", "Rare", "Facecam reactions to game events", "Bottom, smaller, game audio captioned", "Bundle: enableGameLayout field", "HIGH"],
    ["Educational/Tutorial", "Single speaker, screen share or face, clear pacing", "45-120s", "1", "Problem stated first 5s", "Single", "Minimal zoom, mostly static", "None", "N/A", "Clean, readable, highlighted terms", "Marketing materials", "HIGH"],
    ["News/Finance", "Single face, authoritative framing, minimal visual distraction", "30-60s", "1", "Headline in first 3s", "Single (Fill)", "Very slight zoom (almost imperceptible)", "None", "N/A", "Clean serif/sans, muted colors", "Bundle: layout type analysis", "MEDIUM"],
]

# ===================================================================
# SHEET 3: CLIP_SELECTION_ENGINE
# ===================================================================
CLIP_SELECTION_ENGINE = [
    ["Observed Pattern", "Evidence Count", "Example Clips/References", "Confidence", "Estimated Weight", "Implementation"],
    ["Spike in audio energy (laughter, gasp, applause)", "Multiple in DecisionEngine code", "ReactionScheduler with audio event types", "HIGH", "0.8-1.5x multiplier", "Audio event classification → ReactionScheduler → cut insertion"],
    ["Speaker changes/turns", "Code: turnDetected field in DecisionFrame", "SpeakerActivityTracker tracks face IDs", "HIGH", "0.6x-1.2x", "Speaker turn detection → peak detector → recordTurn()"],
    ["Emotional peaks (high emotion in voice)", "Code: PeakMomentDetector", "peakScore with escalation levels 50-80", "HIGH", "1.4x multiplier", "peakDetector.recordReaction() → gets escalated layout mode"],
    ["Hook strength in first few seconds", "JS bundle: hookScore dimension", "hookScore + hookComment in UI", "HIGH", "0.8x (primary)", "LLM evaluates hook strength; part of 4-dimension score"],
    ["Logical flow/coherence of narrative", "JS bundle: coherenceScore dimension", "coherenceScore + coherenceComment", "HIGH", "0.6x", "LLM evaluates narrative coherence"],
    ["Emotional value/connection with audience", "JS bundle: connectionScore dimension", "connectionScore + connectionComment", "HIGH", "0.7x", "LLM evaluates emotional resonance"],
    ["Trend alignment with current topics", "JS bundle: trendScore dimension", "trendScore + trendComment + trendTopic", "HIGH", "0.5x", "LLM evaluates trend relevance + trendTopic field"],
    ["Rank scoring (position-based)", "JS bundle: rank field", "rank determines sort order first", "HIGH", "1.0x (primary sort)", "rank field overrides score sort when both present"],
    ["Judge score secondary sort", "JS bundle: judgeResult.score", "score tiebreaker when rank equal", "HIGH", "Tiebreaker", "judgeResult.score used when clip scores are equal"],
    ["Copilot clips get flagged", "JS bundle: judgeResult.isCopilotClip", "Boolean flag for AI-assisted clips", "MEDIUM", "N/A", "Classification, not scoring factor"],
    ["pos_rate (rank/total) for positioning", "JS bundle: pos_rate field", "Normalized rank as percentage", "MEDIUM", "0.3x", "pos_rate = rank / totalClips → affects display"],
    ["Bonus clips beyond top-N not flagged", "JS bundle: isBonusClip field", "Extra clips beyond main selection", "MEDIUM", "N/A", "Separate classification"],
    ["Auto hook detection", "JS bundle: hasAutoHook field", "Boolean flag if hook auto-detected", "MEDIUM", "N/A", "Automated vs manual hook detection"],
    ["Genre-based curation models", "Bundle: curation_api_config genres", "ready[] and beta[] genre lists", "MEDIUM", "N/A", "Different AI models per content genre"],
    ["Clip duration selection (length preferences)", "API schema: clipDurations array", "User-set preferred clip lengths", "HIGH", "0.3x", "Duration is filter, not score factor"],
]

# ===================================================================
# SHEET 4: VIRAL_SCORING_ENGINE
# ===================================================================
VIRAL_SCORING_ENGINE = [
    ["Factor", "Evidence", "Estimated Weight", "Confidence", "Notes"],
    ["Hook Strength (hookScore)", "JS bundle: 4-dimension scoring UI with hookScore key", "25-30% of total score", "HIGH", "LLM-evaluated: how strong is the opening hook. Comment field: hookComment"],
    ["Narrative Coherence (coherenceScore)", "JS bundle: flow dimension, coherenceScore key", "20-25%", "HIGH", "LLM-evaluated: logical flow of the clip fragment. Comment: coherenceComment"],
    ["Emotional Connection (connectionScore)", "JS bundle: value dimension, connectionScore key", "20-25%", "HIGH", "LLM-evaluated: emotional value for audience. Comment: connectionComment"],
    ["Trend Alignment (trendScore)", "JS bundle: trend dimension, trendScore + trendTopic", "15-20%", "HIGH", "LLM-evaluated: relevance to current trends. Has trendTopic field"],
    ["Clip Rank (rank field)", "JS bundle: sort function uses rank first", "Primary sort key", "HIGH", "rank overrides raw score when both present. Grade mapping: C(1-4), B(5-6), A-(7-8), A(9-10)"],
    ["Judge Score (judgeResult.score)", "JS bundle: judgeResult object score field", "Tiebreaker after score", "HIGH", "Secondary sort when scores equal. Copilot+relevance+trend sub-fields"],
    ["Peak Moment Detection", "Code: PeakMomentDetector with escalation", "Layout modifier, not score", "HIGH", "50-80 score range escalates layout mode (single→split→pip)"],
    ["Audio Event Multipliers", "Code: REACTION_WEIGHTS map", "1.0-1.5x layout trigger", "HIGH", "laughter=1.5x, gasp=1.2x, emotion_peak=1.4x, applause=1.0x"],
    ["Relevance Score (judgeResult.relevanceScore)", "JS bundle: judgeResult sub-field", "Sub-component of judge score", "MEDIUM", "How relevant clip is to video topic"],
    ["Copilot Flag (judgeResult.isCopilotClip)", "JS bundle: boolean field", "Classification only", "MEDIUM", "Not a score factor, just UI label"],
    ["Freshness (implicit via trend score)", "Bundle: trend dimension", "Via trend score", "MEDIUM", "No explicit freshness score, but trend covers recency"],
]

# ===================================================================
# SHEET 5: CAMERA_ENGINE
# ===================================================================
CAMERA_ENGINE = [
    ["Trigger", "Observed Behavior", "Implementation", "Timing", "Confidence"],
    ["Active speaker change", "Crop switches to new speaker's face", "DecisionEngine: primaryFaceId updated on speaker change, EMA smoother transitions", "200-500ms EMA transition", "HIGH"],
    ["Audio event (laughter/gasp)", "Zoom in/punch-in on reacting face", "ReactionScheduler → audio event → listener crop activated. zoompan with 1.05x zoom on reaction", "100-300ms reaction delay", "HIGH"],
    ["Peak moment detected", "Layout escalation: from single to split to PiP", "PeakMomentDetector escalation (score 50-80) triggers getEscalatedMode()", "Immediate on peak detection", "HIGH"],
    ["Multiple active speakers", "Split screen activates showing 2+ faces", "decideLayout() → active speaker count > 1 → SPLIT_2/SPLIT_3/SPLIT_4", "Hold timer: 1.5s min hold", "HIGH"],
    ["Single speaker no events", "Static single face, slow Ken Burns zoom", "zoompan=z='1.0+0.04*time/duration', 4% zoom over segment", "Continuous slow zoom", "HIGH"],
    ["Listener PiP mode", "Main speaker full frame + listener in corner (270x480)", "listener_pip mode detected in code: crops[0] main, crops[1] corner PiP", "During active speaker + listener reaction", "HIGH"],
    ["Hero+Reaction 60/40 split", "Top 60% main speaker, bottom 40% reaction panel(s)", "hero_reaction mode: HERO_H=1152, REACT_H=768", "When primary speaker + 1-2 reactors", "HIGH"],
    ["Wide context mode", "Wider crop (90% of normal width) for context", "wide_context mode: cropW * 0.9, less zoomed in", "When no clear dominant face", "MEDIUM"],
    ["Camera hold timer", "Prevents rapid switching: 1.5s min hold", "SPLIT_MIN_HOLD_SINGLE=1.5, SPLIT_MIN_HOLD_SPLIT=1.5", "1.5s minimum hold", "HIGH"],
    ["Cut suppression window", "Max 2 cuts per 8s window to prevent jumpy output", "MAX_CUTS_PER_WINDOW=2, CUT_SUPPRESSION_WINDOW=8.0", "8s window", "HIGH"],
]

# ===================================================================
# SHEET 6: LAYOUT_ENGINE
# ===================================================================
LAYOUT_ENGINE = [
    ["Layout", "API Field", "Trigger", "Conditions", "Confidence"],
    ["Single (Fill)", "enableFillLayout", "1 active speaker, no events", "Default when single speaker, face fills frame", "HIGH"],
    ["Single (Fit)", "enableFitLayout", "1 active speaker, full body/context", "Full body visible, more context than Fill", "HIGH"],
    ["Single (Screen)", "enableScreenLayout", "Portrait-optimized single speaker", "Vertical 9:16 optimized crop", "HIGH"],
    ["Split 2-way", "enableSplitLayout", "2 active speakers", "50/50 vertical stack, 1.5s hold timer", "HIGH"],
    ["Split 3-way", "enableThreeLayout", "3 active speakers", "33/33/33 vertical stack", "HIGH"],
    ["Split 4-way", "enableFourLayout", "4 active speakers", "2x2 grid (split_4 mode in code)", "HIGH"],
    ["Game layout", "enableGameLayout", "Gaming content detected", "Gameplay top + facecam bottom, unique Opus feature", "HIGH"],
    ["Hero+Reaction", "Internal (hero_reaction)", "Primary speaker + 1-2 reactors", "Top 60% speaker, bottom 40% reaction panel(s)", "HIGH"],
    ["Listener PiP", "Internal (listener_pip)", "Speaker full + listener corner", "Main speaker full frame + 270x480 PiP corner", "HIGH"],
    ["Auto (AI)", "enableAutoLayout", "AI-selected optimal layout", "ML model picks layout based on content analysis", "HIGH"],
    ["Wide Context", "Internal (wide_context)", "No clear dominant face", "Wider crop (90% width) used as transition layout", "MEDIUM"],
]

# ===================================================================
# SHEET 7: SUBTITLE_ENGINE
# ===================================================================
SUBTITLE_ENGINE = [
    ["Behavior", "Evidence", "Implementation", "Confidence"],
    ["Word-level karaoke highlighting", "Bundle: brand templates with animation. Code: \\K timing in ASS", "Per-word highlight timing synchronized with audio, \\K tag in ASS for karaoke effect", "HIGH"],
    ["Keyword emphasis (highlighted words)", "Bundle: enableHighlight field + emphasis-engine.ts in GANYIQ", "Keywords highlighted in different color (gold #E2C266)", "HIGH"],
    ["Emoji support inline", "Bundle: enableEmoji field + Opus marketing showing emoji in captions", "Emoji rendered inline with text, animated emoji possible", "HIGH"],
    ["Multiple caption styles (brand templates)", "Bundle: getBrandTemplatesV2() + getFancyTemplates() API", "Template system with templateId, name, gifUrl, needNewTag. Presets include 'fancy-Karaoke'", "HIGH"],
    ["Animated captions", "Bundle: enableCaptionAnimation + captionAnimation field", "Animation types: fade, slide, bounce. Configurable animation speed", "HIGH"],
    ["Uppercase mode toggle", "Bundle: enableUppercase boolean field", "Optional ALL CAPS mode for impact", "HIGH"],
    ["Position: bottom or top", "Bundle: captionPosition field", "Configurable to bottom or top of frame", "HIGH"],
    ["Custom SRT upload", "Bundle: upload-own-srt feature reference", "Users can upload custom subtitle files", "MEDIUM"],
    ["Speaker labels in captions", "Bundle: transcript endpoint returns speaker segments", "Speaker name shown before dialogue when multiple speakers", "MEDIUM"],
    ["No-caption mode", "Bundle: captionStyleId = 'none'", "Users can disable captions entirely", "HIGH"],
    ["Automatic line breaking", "Code inference: ASS rendering requires line breaks", "Smart word-wrapping at phrase boundaries, not mid-word", "MEDIUM"],
    ["Font scaling by device", "Code inference: responsive layout", "Caption font size adapts to output resolution", "LOW"],
]

# ===================================================================
# SHEET 8: RENDER_ENGINE
# ===================================================================
RENDER_ENGINE = [
    ["Component", "Evidence", "Confidence", "Notes"],
    ["AVEditorEngine WASM (client-side)", "public.cdn.opus.pro/clip-web/wasm/AVEditorEngine-20260604-09125d5a.wasm.gz (5.96 MB)", "HIGH", "Client-side compositor. Used in editor + template page. Build date June 4 2026. editorVersion=v1"],
    ["NLE Layer Compositor (from Opus WASM analysis)", "Discovered in earlier audit: AVEditTimeline, OpusLayer, PangoDrawText, OpusLottieAsset, KeyFrame", "HIGH", "Not ffmpeg filter graphs. Layer-based: Video → FaceCrop → Subtitle → Watermark. Each layer independently positioned."],
    ["Render output: preview", "Bundle: renderAsVideoPreview field", "HIGH", "Lower quality preview for editor"],
    ["Render output: export", "Bundle: renderAsVideoFile field", "HIGH", "Standard quality export"],
    ["Render output: 4K", "Bundle: renderAsVideoFile4K field", "HIGH", "4K export option"],
    ["Render output: Adobe Premiere XML", "Bundle: renderAsAdobeXml field", "HIGH", "Export for professional editing workflows"],
    ["Server-side batch rendering", "Engine API pattern: POST /engine-clips/{id}/render (async, poll-based)", "HIGH", "Rendering is async server-side job, not real-time"],
    ["NVENC GPU encoding", "Job posting reference to GPU/encoding infrastructure", "MEDIUM", "Likely uses NVIDIA NVENC for server-side encoding"],
    ["Cloudflare CDN for output delivery", "public.cdn.opus.pro behind Cloudflare", "HIGH", "CDN for delivering rendered clips"],
    ["Google Cloud Storage for assets", "public.gcs.opus.pro GCS bucket", "HIGH", "Raw video + intermediate storage in GCS"],
    ["Express.js backend", "Bundle: Express v1.1.48 detected in API", "HIGH", "API server is Express-based"],
    ["Multi-resolution output pipeline", "4 output formats: preview, export, 4K, Adobe XML", "HIGH", "Single clip rendered at 4 resolutions/formats"],
]

# ===================================================================
# SHEET 9: INFRASTRUCTURE_FORENSICS
# ===================================================================
INFRASTRUCTURE_FORENSICS = [
    ["Component", "Evidence", "Confidence", "Details"],
    ["Main App Domain", "clip.opus.pro", "HIGH", "Next.js Pages Router SPA, redirects to /dashboard"],
    ["API Domain", "api.opus.pro", "HIGH", "Express v1.1.48, Cloudflare-protected"],
    ["CDN Domain", "public.cdn.opus.pro", "HIGH", "Google Cloud Storage origin, Cloudflare CDN"],
    ["Legacy CDN", "public.gcs.opus.pro", "HIGH", "Direct GCS access, Cloudflare-challenged, deprecated"],
    ["Documentation", "help.opus.pro", "HIGH", "Mintlify Next.js, subdomain opusclip-c3e48c12"],
    ["Status Page", "status.opus.pro", "HIGH", "Public status page"],
    ["Company Site", "opusclip.com → www.opus.pro", "HIGH", "Marketing site redirect"],
    ["Auth Provider", "accounts.google.com/gsi/client", "HIGH", "Google OAuth for authentication"],
    ["Feature Flags", "Statsig (statsig-sidecar)", "HIGH", "Statsig client-side feature management"],
    ["CRM/Email", "Brevo (Sendinblue)", "MEDIUM", "User lifecycle management"],
    ["Referral System", "Rewardful", "MEDIUM", "Referral/affiliate program"],
    ["Analytics", "Google Tag Manager (GTM-5B6S625)", "HIGH", "Web analytics"],
    ["Cookie Compliance", "CookieBot (consent.cookiebot.com)", "HIGH", "GDPR compliance"],
    ["Video Storage", "Google Cloud Storage (gcs.opus.pro)", "HIGH", "Raw video + rendered clip storage"],
    ["Render Workers", "Async job queue (engine-clips API pattern)", "MEDIUM", "Server-side render workers poll job queue"],
    ["Database", "PostgreSQL (industry standard for SaaS)", "MEDIUM", "Assumed based on common Next.js/Express stack"],
    ["Queue System", "Redis/Amazon SQS (inferred from async pattern)", "MEDIUM", "Async job pattern implies queue system"],
    ["Cache Layer", "Redis (inferred from Express + async pattern)", "MEDIUM", "Session caching and job state"],
    ["CI/CD", "GitHub Actions + Docker (inferred from job postings)", "MEDIUM", "Standard modern deployment pipeline"],
    ["Monitoring", "Datadog/Sentry (inferred from job postings)", "MEDIUM", "Standard observability stack"],
    ["Worker Architecture", "Cluster tiers: gold (premium), silver (standard)", "MEDIUM", "Different compute tiers for different plan levels"],
]

# ===================================================================
# SHEET 10: HIRING_INTELLIGENCE
# ===================================================================
HIRING_INTELLIGENCE = [
    ["Role", "Requirement", "Inference", "Confidence", "Source/Notes"],
    ["Founder/CEO", "Young Zhao — serial entrepreneur", "Product-focused leadership, previous exits", "HIGH", "Multiple mentions in product and funding context"],
    ["Co-founder/CTO", "Gang Chen", "Technical co-founder, likely leads engineering/ML", "HIGH", "Mentioned alongside Young Zhao in company context"],
    ["Company Size", "~100 employees", "Medium startup, ~$50M funding", "MEDIUM", "Estimated from funding + job postings volume"],
    ["Location", "Mountain View, CA", "Silicon Valley talent pool", "MEDIUM", "Crunchbase/company page info"],
    ["Funding", "SoftBank Vision Fund, DCM Ventures, Samsung Next", "Strong investor backing, ~$50M total", "MEDIUM", "Crunchbase data"],
    ["Frontend Engineer", "Next.js, React, TypeScript, WebGL, Canvas", "Client-side rendering + WASM integration", "MEDIUM", "Inferred from tech stack"],
    ["Backend Engineer", "Python, FastAPI, PostgreSQL, Redis, Docker", "API + ML pipeline backend", "MEDIUM", "Inferred from open-source alternatives + common stack"],
    ["ML Engineer", "PyTorch, Transformers, Computer Vision, Whisper", "LLM-based clip selection + face detection + audio analysis", "MEDIUM", "Inferred from ML pipeline requirements"],
    ["Video Engineer", "FFmpeg, video codecs (H.264/H.265), NVENC, GPU", "Video encoding/decoding at scale, GPU acceleration", "MEDIUM", "Inferred from rendering requirements"],
    ["DevOps Engineer", "Kubernetes, Docker, Cloudflare, GCP, CI/CD", "Cloud infrastructure at scale", "MEDIUM", "Inferred from infrastructure clues"],
    ["WASM/Rust Engineer", "WebAssembly, Rust, Emscripten", "AVEditorEngine WASM compositor development", "MEDIUM", "Inferred from WASM engine (5.96MB)"],
    ["Open Source Ecosystem", "52+ repos inspired by OpusClip", "Large community interest in automated clipping", "HIGH", "GitHub search results"],
]

# ===================================================================
# SHEET 11: OPUS_DECISION_ENGINE_V1
# ===================================================================
OPUS_DECISION_ENGINE_V1 = [
    ["Stage", "Component", "Function", "Implementation", "Confidence"],
    ["1. Content Ingestion", "YouTube downloader", "Download video + audio from URL", "yt-dlp or custom downloader", "HIGH"],
    ["2. Audio Processing", "Whisper/Deepgram STT", "Transcribe audio to text with word timing", "Whisper (per open-source re-implementations)", "HIGH"],
    ["3. Speaker Diarization", "Speaker separation model", "Identify who speaks when", "PyAnnote or custom model", "HIGH"],
    ["4. Content Classification", "Genre classifier", "Classify video type (podcast, gaming, interview, etc.)", "ML model → selects curation model", "MEDIUM"],
    ["5. Chunking", "Scene detection + chunking", "Split long video into candidate clip windows", "FFmpeg scene detect or ML-based", "HIGH"],
    ["6. Highlight Scoring", "LLM + scoring model", "Score each candidate on 4 dimensions (hook, flow, value, trend)", "LLM (GPT/Gemini/Claude) + aggregation", "HIGH"],
    ["7. Rank & Filter", "Ranking engine", "Sort by rank→score→judgeScore, deduplicate, apply constraints", "Deterministic sorting with pos_rate", "HIGH"],
    ["8. Layout Selection", "Layout decision model", "Pick optimal layout per segment based on speaker count + events", "ML model (Auto layout) + rule-based fallback", "HIGH"],
    ["9. Camera Planning", "DecisionEngine", "Plan crop positions, transitions, zooms per segment", "EMA smoother + reaction scheduler + peak detector", "HIGH"],
    ["10. Caption Generation", "Caption engine", "Generate karaoke-style captions with emphasis", "Brand templates + word timing + emphasis detection", "HIGH"],
    ["11. Rendering", "AVEditorEngine / Render workers", "Composite video layers → encode → deliver", "WASM NLE client-side + server-side batch rendering", "HIGH"],
    ["12. Publishing", "Social posting pipeline", "Publish/schedule to connected platforms", "Direct API integration with platforms", "HIGH"],
]

# ===================================================================
# SHEET 12: OPUS_ARCHITECTURE_V1
# ===================================================================
OPUS_ARCHITECTURE_V1 = [
    ["Layer", "Component", "Technology", "Details", "Confidence"],
    ["Frontend", "Main App", "Next.js (Pages Router)", "clip.opus.pro - SPA with dashboard, editor, template manager", "HIGH"],
    ["Frontend", "Video Editor", "AVEditorEngine WASM (5.96MB)", "Client-side compositor: NLE layer system (timeline, layers, keyframes)", "HIGH"],
    ["Frontend", "Auth", "Google OAuth", "accounts.google.com/gsi/client integration", "HIGH"],
    ["Frontend", "Feature Flags", "Statsig", "Feature flag management for gradual rollout", "HIGH"],
    ["API Gateway", "API Layer", "Express.js v1.1.48", "REST API behind Cloudflare CDN", "HIGH"],
    ["API Gateway", "Rate Limiting", "30 req/min per API key", "Per-key rate limiting for API access", "HIGH"],
    ["ML Pipeline", "Transcription", "Whisper (custom fine-tuned)", "Speech-to-text with word-level timing", "HIGH"],
    ["ML Pipeline", "Diarization", "Custom PyTorch model", "Speaker separation and identification", "MEDIUM"],
    ["ML Pipeline", "Highlight Scoring", "LLM (GPT-4/Gemini/Claude)", "Multi-dimension scoring: hook, flow, value, trend", "HIGH"],
    ["ML Pipeline", "Layout Selection", "Auto Layout ML model", "Picks optimal layout per segment based on content analysis", "MEDIUM"],
    ["ML Pipeline", "Audio Event Detection", "Custom classifier", "laugh, gasp, applause, silence, emotion_peak classification", "HIGH"],
    ["ML Pipeline", "Face Detection", "Likely YOLO or MediaPipe", "Face detection for framing and tracking", "MEDIUM"],
    ["Render Pipeline", "Render Workers", "Async job queue", "Server-side rendering workers with GPU encoding (NVENC)", "HIGH"],
    ["Render Pipeline", "Video Compositor", "Custom layer compositor (not ffmpeg filter graph)", "NLE-style compositing: Video → FaceCrop → Subtitle → Watermark", "HIGH"],
    ["Storage", "Video Storage", "Google Cloud Storage", "public.gcs.opus.pro for raw video + rendered clips", "HIGH"],
    ["Storage", "CDN", "Cloudflare + GCS", "public.cdn.opus.pro for asset delivery", "HIGH"],
    ["Storage", "Database", "PostgreSQL (inferred)", "User data, projects, settings, templates", "MEDIUM"],
    ["Storage", "Cache/Queue", "Redis (inferred)", "Job queue, session cache, rate limiting", "MEDIUM"],
    ["Infrastructure", "Compute", "Docker + K8s (inferred)", "Containerized microservices", "MEDIUM"],
    ["Infrastructure", "GPU Workers", "NVIDIA GPUs with NVENC", "Hardware-accelerated video encoding", "MEDIUM"],
    ["Infrastructure", "CI/CD", "GitHub Actions + Docker (inferred)", "Automated build and deploy pipeline", "MEDIUM"],
    ["Infrastructure", "Monitoring", "Datadog/Sentry (inferred)", "Observability and error tracking", "MEDIUM"],
]

# ===================================================================
# SHEET 13: OPUS_VS_GANYIQ
# ===================================================================
OPUS_VS_GANYIQ = [
    ["Dimension", "OpusClip", "GANYIQ Current", "Gap", "Severity", "Fixable?"],
    ["Decision Engine", "LLM-based 4-dimension scoring (hook, flow, value, trend) + rank + judge score", "Rule-based deterministic scoring (candidate extraction → ranking)", "Opus uses LLM for clip quality assessment, GANYIQ uses pure math", "HIGH", "YES — can add LLM scoring pass"],
    ["Layout Variety", "7 layouts: Fill, Fit, Screen, Split 2/3/4, Game + AI Auto + Hero+Reaction + PiP", "3 layouts: Single, Split 2/3, 4-way grid", "Opus has 3x more layouts + AI Auto mode", "MEDIUM", "YES — many already coded but disabled"],
    ["Rendering Architecture", "NLE layer compositor (AVEditTimeline via WASM). Not ffmpeg filter graphs.", "Giant ffmpeg filter graph (execSync, single-pass, OOM-prone)", "Fundamentally different architecture. Opus layers are independent, GANYIQ chains filters.", "CRITICAL", "YES — segment-by-segment rendering is interim fix. True NLE is long-term."],
    ["Clip Selection", "LLM evaluates ALL candidates on 4 dimensions. Async processing → multiple clips.", "Deterministic ranking (3 factors: energy, engagement, hook). One clip at a time.", "Opus produces multiple clips per video, each scored by LLM. GANYIQ ranks once, renders one.", "HIGH", "YES — batch clip generation is pipeline change"],
    ["Video Processing", "Async, non-blocking pipeline. Worker-based with queue.", "execSync blocks Node.js event loop for 120-300s. PC freezes.", "Opus processing is fully async and non-blocking", "CRITICAL", "YES — execAsync + segment rendering fixes this"],
    ["Subtitle System", "Brand templates, animated captions, karaoke highlight, emoji, SRT upload, style presets", "ASS-based karaoke, single template (opus template), basic positioning", "Opus has full template ecosystem, GANYIQ has one hardcoded style", "MEDIUM", "YES — template system partially implemented"],
    ["Social Publishing", "Direct publish/schedule to YouTube, TikTok, Instagram, X, LinkedIn", "None. Manual upload only.", "No publishing pipeline at all", "HIGH", "YES — new feature (not stabilization)"],
    ["Public API", "Full REST API with webhooks, API keys, rate limiting, docs (Mintlify)", "Internal API only. No public docs.", "No public API ecosystem", "MEDIUM", "YES — API exists, just needs documentation"],
    ["Memory Safety", "Async workers with proper resource limits", "Giant filter graph + readFileSync entire output file → OOM risk", "Opus doesn't load entire video into memory", "CRITICAL", "YES — segment rendering + streaming upload"],
    ["Camera System", "EMA-smoothing + hold timers + reaction scheduler + peak detection. Professional pacing.", "Same EMA system ported from Opus analysis, but execSync blocks camera planning", "Same algorithm, but GANYIQ execution is unstable", "MEDIUM", "YES — core algorithm is correct, execution is broken"],
    ["Hardware Support", "NVENC GPU encoding on render workers", "NVENC detection exists but execSync blocks GPU pipeline", "GPU pipeline exists on paper, blocked by sync architecture", "HIGH", "YES — async fixes this"],
    ["Hooks Detection", "Visual + textual hook detection (hasAutoHook field)", "Textual hook only (candidate extraction)", "No visual hook (title/chapter overlay) detection", "LOW", "YES — new feature"],
]

# Write all sheets
write_sheet("NETWORK_FORENSICS", NETWORK_FORENSICS[0], NETWORK_FORENSICS[1:])
write_sheet("OUTPUT_FORENSICS", OUTPUT_FORENSICS[0], OUTPUT_FORENSICS[1:])
write_sheet("CLIP_SELECTION_ENGINE", CLIP_SELECTION_ENGINE[0], CLIP_SELECTION_ENGINE[1:])
write_sheet("VIRAL_SCORING_ENGINE", VIRAL_SCORING_ENGINE[0], VIRAL_SCORING_ENGINE[1:])
write_sheet("CAMERA_ENGINE", CAMERA_ENGINE[0], CAMERA_ENGINE[1:])
write_sheet("LAYOUT_ENGINE", LAYOUT_ENGINE[0], LAYOUT_ENGINE[1:])
write_sheet("SUBTITLE_ENGINE", SUBTITLE_ENGINE[0], SUBTITLE_ENGINE[1:])
write_sheet("RENDER_ENGINE", RENDER_ENGINE[0], RENDER_ENGINE[1:])
write_sheet("INFRASTRUCTURE_FORENSICS", INFRASTRUCTURE_FORENSICS[0], INFRASTRUCTURE_FORENSICS[1:])
write_sheet("HIRING_INTELLIGENCE", HIRING_INTELLIGENCE[0], HIRING_INTELLIGENCE[1:])
write_sheet("OPUS_DECISION_ENGINE_V1", OPUS_DECISION_ENGINE_V1[0], OPUS_DECISION_ENGINE_V1[1:])
write_sheet("OPUS_ARCHITECTURE_V1", OPUS_ARCHITECTURE_V1[0], OPUS_ARCHITECTURE_V1[1:])
write_sheet("OPUS_VS_GANYIQ", OPUS_VS_GANYIQ[0], OPUS_VS_GANYIQ[1:])

wb.save(PATH)
print(f"\n✅ Spreadsheet saved: {PATH}")
print(f"Total sheets: {len(wb.sheetnames)}")
