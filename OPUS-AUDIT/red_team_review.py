#!/usr/bin/env python3
"""OPUS AUDIT RED TEAM REVIEW — Critical self-audit of all claims"""

import openpyxl
from openpyxl.styles import Font, PatternFill

PATH = '/root/GANYIQ/OPUS-AUDIT/opus_audit_master.xlsx'
wb = openpyxl.load_workbook(PATH)

HEADER_FILL = PatternFill(start_color='1F1F2E', end_color='1F1F2E', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
ALT_FILL = PatternFill(start_color='F5F5FF', end_color='F5F5FF', fill_type='solid')
RED_FILL = PatternFill(start_color='FFDDDD', end_color='FFDDDD', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFFFDD', end_color='FFFFDD', fill_type='solid')
GREEN_FILL = PatternFill(start_color='DDFFDD', end_color='DDFFDD', fill_type='solid')

def style_header(ws, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

def write_sheet(name, headers, rows):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(title=name)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=str(val))
        if r % 2 == 0:
            for c in range(1, len(headers)+1):
                ws.cell(row=r, column=c).fill = ALT_FILL
    style_header(ws, len(headers))
    ws.auto_filter.ref = f"A1:{chr(64+len(headers))}{len(rows)+1}"
    print(f"  {name}: {len(rows)} rows")

# ===================================================================
# SHEET 1: CLAIM_REVIEW — Every major claim + brutal reassessment
# ===================================================================
CLAIM_REVIEW = [
    ["Claim", "Previous Classification", "Previous Confidence", "Reassessment", "New Classification", "New Confidence", "Rationale for Downgrade/Upgrade", "Evidence Gap"],
    
    ("Opus uses LLM-powered scoring", "INFERENCE", "MEDIUM",
     "Overclaimed — official docs say 'AI evaluates'. AI != LLM. Could be BERT-based classifier or smaller model. No LLM API calls visible in any bundle.",
     "INFERENCE", "LOW",
     "Docs say 'AI' not 'LLM'. Bundle has 0 LLM API call references. 'LLM' is assumption.",
     "Need server-side evidence of actual model type"),
    
    ("Opus uses 4-dimension scoring (Hook + Flow + Value + Trend)", "FACT", "HIGH",
     "Partially confirmed — docs confirm Hook + Flow + Value (3 dims). Trend ONLY in JS bundle, NOT in official docs. Overclaimed as 4 confirmed dimensions.",
     "FACT/INFERENCE", "MEDIUM",
     "3 dims = FACT (official docs). 4th dim (Trend) = INFERENCE (bundle only). Should not be presented as 4 confirmed dimensions.",
     "Need official documentation of Trend dimension or weight"),
    
    ("Scoring weights: Hook 25-30%, Flow 20-25%, Value 20-25%, Trend 15-20%", "INFERENCE", "MEDIUM",
     "COMPLETE GUESSWORK. No evidence of weights anywhere. Presenting percentages implies precision we don't have.",
     "SPECULATION", "LOW",
     "Zero evidence for any weight values. Made up based on 'seems important'. This is guessing.",
     "Need ANY evidence of actual scoring weights"),
    
    ("Audio event multipliers (laughter=1.5x, gasp=1.2x, etc.)", "INFERENCE", "MEDIUM",
     "GANYIQ values presented as Opus values. These are from GANYIQ's port, not from Opus code. Actual Opus multipliers unknown.",
     "SPECULATION", "LOW",
     "Presenting GANYIQ's interpreted values as Opus fact. No evidence these are Opus values.",
     "Need Opus server-side code for actual multipliers"),
    
    ("Opus uses NLE layer compositor (AVEditTimeline, OpusLayer, PangoDrawText)", "INFERENCE", "MEDIUM",
     "Based on PREVIOUS WASM binary analysis that hasn't been re-verified. Symbols could be from embedded third-party libraries. 'AVEditTimeline' sounds like a custom class but 'PangoDrawText' is a known open-source text layout library.",
     "INFERENCE", "LOW",
     "Haven't re-verified from current WASM (5.96MB built June 4 2026). Previous analysis may be from different version. Pango is open-source text library, not unique Opus compositor.",
     "Need fresh WASM binary analysis of current version"),
    
    ("Opus does NOT use ffmpeg", "FACT", "HIGH",
     "OVERCLAIMED. Zero ffmpeg in frontend bundles PROVES frontend doesn't use ffmpeg. Does NOT prove server-side rendering doesn't use ffmpeg. Server-side encoding almost certainly uses ffmpeg or similar.",
     "FACT (frontend) / SPECULATION (server)", "LOW",
     "Frontend = FACT. Server-side = pure speculation. In a modern video stack, ffmpeg is used somewhere for encoding. The claim 'does not use ffmpeg' is too broad.",
     "Need server-side infrastructure access"),
    
    ("Opus uses batch processing (multiple clips per video)", "FACT", "HIGH",
     "CONFIRMED — API creates projects with multiple clips. Each project returns array of clips with scores and ranks. This is the strongest evidence for batch processing.",
     "FACT", "HIGH",
     "Direct API evidence + clip data model supporting multiple clips per project. No inference needed.",
     "None — well supported"),
    
    ("Opus uses GPU encoding / NVENC", "SPECULATION", "LOW",
     "Correctly classified as speculation. Zero evidence for GPU/NVENC anywhere.",
     "SPECULATION", "LOW",
     "No changes needed — original classification was accurate.",
     "Need infrastructure evidence"),
    
    ("Opus ranking: rank > score > judgeResult.score", "FACT", "HIGH",
     "CONFIRMED. Sort function literal found in JS bundle. Code shows exact sort order with rank, score, and judgeResult.score tiebreaker.",
     "FACT", "HIGH",
     "Direct evidence from JS bundle. Sort function literal is unambiguous.",
     "None — well supported"),
    
    ("Grade mapping: C(≤4), B(5-6), A-(7-8), A(9-10)", "FACT", "HIGH",
     "CONFIRMED. Grade function literal found. But this is UI-ONLY. Not backend scoring logic.",
     "FACT (UI) / SPECULATION (backend)", "MEDIUM",
     "Function literal is clear but applies to UI display only. Backend may use completely different grading.",
     "Unknown if backend uses same grades"),
    
    ("Opus uses ClipAnything for natural language retrieval", "FACT", "HIGH",
     "CONFIRMED. Official help.opus.pro documentation explicitly describes ClipAnything feature. 6 indexed categories documented.",
     "FACT", "HIGH",
     "Official documentation is unambiguous. Feature description is detailed.",
     "None — well supported"),
    
    ("Opus virality score = 0-99 based on Hook+Flow+Value", "FACT", "HIGH",
     "CONFIRMED. Official virality score documentation explicitly states range 0-99 and evaluation dimensions.",
     "FACT", "HIGH",
     "Direct from official docs. Score range, dimensions, and behavior all documented.",
     "None — well supported"),
    
    ("Engine clips API (save/render/refine pattern)", "FACT", "HIGH",
     "CONFIRMED. Found all 5 API endpoints in dashboard.js bundle.",
     "FACT", "HIGH",
     "Endpoint strings confirmed in JS bundle.",
     "None — well supported"),
    
    ("Social publishing API (post, schedule, cancel, copy)", "FACT", "HIGH",
     "CONFIRMED. Multiple endpoints found in bundle + official docs.",
     "FACT", "HIGH",
     "Endpoint documentation + bundle references.",
     "None — well supported"),
    
    ("EMA smoothing for camera transitions", "INFERENCE", "MEDIUM",
     "IMPORTANT NOTE: This claim is based on GANYIQ's codebase (ported from Opus behavior analysis). The actual Opus implementation may be completely different. We inferred this from OBSERVING Opus output, not from Opus code.",
     "INFERENCE", "LOW",
     "GANYIQ implementation = GANYIQ's interpretation of Opus behavior. Actual Opus algorithm unknown.",
     "Need Opus code for camera decision logic"),
    
    ("Hold timers (1.5s), Cut suppression (2 cuts/8s), Peak escalation (50-80)", "INFERENCE", "MEDIUM",
     "Same issue as EMA — these are GANYIQ's ported values, tuned for GANYIQ output. Presenting them as Opus values is misleading.",
     "SPECULATION", "LOW",
     "Numerical thresholds are GANYIQ's tuning, not Opus values. Presenting them as discovered Opus parameters is incorrect.",
     "Need Opus code or behavior analysis for actual thresholds"),
    
    ("7 layout types + Auto + Game", "FACT", "HIGH",
     "CONFIRMED. 8 enableLayout flags found in both API docs and JS bundle. Game layout is unique.",
     "FACT", "HIGH",
     "Direct API documentation + bundle references.",
     "None — well supported"),
    
    ("Brand template system for captions", "FACT", "HIGH",
     "CONFIRMED. getBrandTemplatesV2(), getFancyTemplates() documented. templateId, name, gifUrl confirmed.",
     "FACT", "HIGH",
     "Direct evidence from bundle API calls + official docs.",
     "None — well supported"),
    
    ("PostgreSQL database", "INFERENCE", "MEDIUM",
     "This is WAG. No evidence for any specific database. Could be MySQL, MongoDB, DynamoDB, etc.",
     "SPECULATION", "LOW",
     "No evidence whatsoever. Common SaaS stack assumption.",
     "Need ANY database evidence"),
    
    ("Redis job queue", "INFERENCE", "MEDIUM",
     "Same as PostgreSQL — no evidence. Could be SQS, RabbitMQ, Bull, Sidekiq, etc.",
     "SPECULATION", "LOW",
     "No evidence. Async pattern implies queue but doesn't specify technology.",
     "Need infrastructure evidence"),
    
    ("Company ~100 employees, $50M funding", "FACT", "MEDIUM",
     "From web research — reasonably confirmed but sources may be outdated.",
     "FACT", "MEDIUM",
     "Cross-referenced from multiple sources. Acceptable.",
     "None — acceptable"),
    
    ("Engineers require Python/FastAPI, FFmpeg, Rust, CUDA", "SPECULATION", "LOW",
     "No live job postings could be fetched (Firecrawl credits exhausted, browser unavailable). These are assumptions from open-source ecosystem.",
     "SPECULATION", "LOW",
     "No live data collected. Tool limitations prevented verification.",
     "Need live job posting access"),
]

# ===================================================================
# SHEET 2: EVIDENCE_CHAIN_REVIEW
# ===================================================================
EVIDENCE_CHAIN_REVIEW = [
    ["Claim", "Evidence Chain", "Logical Leap", "Gap Size", "Assessment"],
    
    ("4-dimension LLM scoring",
     "Official docs: Hook+Flow+Value → JS Bundle: 4 scoreKeys → Conclusion: 4-dim LLM",
     "1. Docs say 'AI' not 'LLM'. 2. 4th dimension (Trend) only in bundle, not docs. 3. Score keys are in frontend UI config, not backend scoring code.",
     "MEDIUM",
     "3 dimensions = solid. LLM claim = leap. Trend dimension = unconfirmed. Need backend evidence."),
    
    ("Scoring weights 25-30%, 20-25%, etc",
     "Hook+Flow+Value described in docs → Conclusion: weights are roughly equal → Assigned numbers",
     "No evidence for any weight values. Assigning percentages is pure fabrication.",
     "CRITICAL",
     "Complete guesswork. Remove weight estimates or mark clearly as speculation."),
    
    ("Opus does NOT use ffmpeg",
     "Zero ffmpeg hits in frontend JS → Conclusion: Opus doesn't use ffmpeg",
     "Frontend ≠ backend. Server-side rendering almost certainly uses ffmpeg or similar for encoding step.",
     "CRITICAL",
     "Overbroad conclusion from narrow evidence. Frontend = confirmed. Server-side = unknown."),
    
    ("NLE layer compositor",
     "Previous WASM analysis found AVEditTimeline, OpusLayer → Conclusion: NLE compositor",
     "1. WASM not re-verified. 2. Symbols may be from embedded libraries. 3. 'Timeline' could be UI component, not render compositor.",
     "MEDIUM",
     "Needs fresh WASM verification. Current classification (INFERENCE/MEDIUM) is fair but not verified against latest WASM."),
    
    ("EMA smoothing + hold timers + cut suppression",
     "Observed Opus output behavior → Ported algorithm to GANYIQ → Conclusion: Opus uses these exact parameters",
     "1. Observing output ≠ reverse engineering code. 2. GANYIQ's values were tuned by GANYIQ dev, not extracted from Opus.",
     "LARGE",
     "Pattern likely correct (smooth transitions, hold stability) but exact numerical thresholds are GANYIQ guesses."),
    
    ("Audio event multipliers 1.0-1.5x",
     "Observed Opus reaction behavior → GANYIQ implemented ReactionScheduler with weights → Conclusion: Opus uses these weights",
     "GANYIQ weights may be completely different from Opus weights.",
     "LARGE",
     "Only the EXISTENCE of audio event detection is likely. Exact multipliers are GANYIQ's interpretation."),
    
    ("Batch processing architecture",
     "API returns multiple clips per project → Conclusion: Server-side batch processing",
     "Multiple clips could be rendered sequentially, not in parallel. 'Batch' implies parallel which is unproven.",
     "SMALL",
     "Multi-clip output = confirmed. Parallel vs sequential processing = unknown. Claim 'batch' may overstate."),
    
    ("Clip selection ranking: rank > score > judgeScore",
     "Sort function literal found in JS bundle → Conclusion: This is the production ranking",
     "Sort function is client-side for display. Backend may use different sort for actual clip selection.",
     "SMALL",
     "Function confirmed in bundle. May be display-only. Backend may pre-sort differently."),
    
    ("ClipAnything 6 search categories",
     "Official docs: 'scenes, actions, characters, events, emotional moments, viral topics'",
     "No leap — direct evidence from official documentation.",
     "NONE",
     "Solid claim. Well supported by official docs."),
    
    ("PostgreSQL database",
     "Common SaaS stack → Conclusion: Opus uses PostgreSQL",
     "No evidence whatsoever. 'Common stack' is not evidence.",
     "CRITICAL",
     "Remove or change to SPECULATION/LOW."),
    
    ("Redis job queue",
     "Async job pattern → Conclusion: Redis",
     "Async pattern → queue is valid inference. Redis specifically is guesswork.",
     "LARGE",
     "Queue architecture = valid inference. Redis specifically = speculation."),
]

# ===================================================================
# SHEET 3: DECISION_ENGINE_CONFIDENCE
# ===================================================================
DECISION_ENGINE_CONFIDENCE = [
    ["Component", "Current Confidence", "Red-Team Confidence", "Reason for Change", "Bottleneck"],
    ("Candidate generation method", "MEDIUM", "LOW-MEDIUM (30%)", "No evidence of exact chunking strategy. Scene detection + transcript paragraph boundaries is best guess.", "Need Opus source or detailed timing analysis"),
    ("Candidate scoring function", "HIGH", "MEDIUM (55%)", "3 dims confirmed. Trend dim inferred. LLM vs ML model unknown. Weight distribution unknown.", "Need backend scoring code"),
    ("Hook scoring", "HIGH", "MEDIUM (50%)", "Dimension confirmed but evaluation method unknown. 'AI evaluates' = black box.", "Need hook scoring implementation"),
    ("Flow/Coherence scoring", "HIGH", "MEDIUM (50%)", "Dimension confirmed. Same issue — evaluation method unknown.", "Need flow scoring implementation"),
    ("Value/Connection scoring", "HIGH", "MEDIUM (50%)", "Dimension confirmed. Same issue.", "Need value scoring implementation"),
    ("Trend scoring", "MEDIUM", "LOW (25%)", "Only in JS bundle, not in official docs. May be part of ClipAnything only, not general scoring.", "Need official docs or backend code"),
    ("Prompt relevance scoring", "HIGH", "MEDIUM (40%)", "Official docs confirm prompt relevance evaluation. But exact mechanism unknown — could be simple cosine similarity, not complex scoring.", "Need ClipAnything implementation details"),
    ("Final ranking & selection", "HIGH", "HIGH (80%)", "Sort function confirmed. Grouping by taskId confirmed. isBonusClip confirmed. pos_rate confirmed.", "Well supported — minor gap on whether backend uses same sort"),
    ("Deduplication logic", "MEDIUM", "LOW-MEDIUM (30%)", "Grouping by taskId implies dedup within groups. Exact dedup strategy unknown.", "Need dedup implementation details"),
    ("Clip post-processing (trim, extend, caption fix)", "MEDIUM", "MEDIUM (50%)", "Edit endpoints confirmed. User can edit clips after generation. But edit capabilities unknown.", "Need API call details"),
]

# ===================================================================
# SHEET 4: UNKNOWNS_REGISTRY
# ===================================================================
UNKNOWNS_REGISTRY = [
    ["#", "Unknown", "Importance", "Current Status", "What We'd Need to Know"],
    (1, "Actual LLM/ML model used for scoring", "CRITICAL", "Inferred (docs say 'AI')", "Model provider, size, fine-tuning approach"),
    (2, "Scoring weight distribution", "CRITICAL", "Guessed (no evidence)", "Actual weights for Hook, Flow, Value, Trend dimensions"),
    (3, "Candidate chunking strategy", "HIGH", "Best guess", "How are candidates generated? Scene-based? Sliding window? Transcript-based?"),
    (4, "Server-side rendering pipeline", "HIGH", "Unknown", "Does server use ffmpeg? Custom encoder? Cloud service?"),
    (5, "Worker architecture", "HIGH", "Unknown", "How many workers? GPU type? Queue system? Auto-scaling?"),
    (6, "Training data for scoring model", "HIGH", "Unknown", "What videos was model trained on? Human-labeled? How many examples?"),
    (7, "Human review / feedback loop", "HIGH", "Unknown", "Do humans review clips? Is user engagement data fed back into model?"),
    (8, "Edit capabilities (trim, extend, fix captions)", "MEDIUM", "Partially known", "Endpoints exist but exact edit operations unknown"),
    (9, "Scoring model update frequency", "MEDIUM", "Unknown", "How often is model retrained? Real-time? Weekly?"),
    (10, "ClipAnything retrieval mechanism", "MEDIUM", "Inferred (docs describe)", "Embedding-based? Keyword-based? Hybrid?"),
    (11, "Genre-specific model differences", "MEDIUM", "Known (genre list exists)", "How do models differ per genre? Same architecture? Different training?"),
    (12, "Database technology", "LOW", "Unknown", "PostgreSQL, MySQL, MongoDB, or something else?"),
    (13, "Queue technology", "LOW", "Unknown", "Redis, SQS, RabbitMQ, Bull?"),
    (14, "Hosting provider", "LOW", "Unknown", "AWS, GCP, Azure, or self-hosted?"),
    (15, "ML infrastructure", "MEDIUM", "Unknown", "GPU type, model serving framework, inference optimization"),
    (16, "API rate limiting implementation", "LOW", "Known (30 req/min per docs)", "Implementation detail only"),
    (17, "Storage strategy (hot/cold tier)", "LOW", "Unknown", "How long are videos stored? Re-encoding vs cached?"),
    (18, "Copilot feature details", "MEDIUM", "Known (isCopilotClip flag exists)", "How does copilot-assisted clipping work?"),
    (19, "Auto-layout selection model", "MEDIUM", "Inferred (ML model)", "Model type, training data, decision criteria"),
    (20, "Real-time feedback / engagement tracking", "MEDIUM", "Unknown", "Does Opus track clip performance and use it for future scoring?"),
]

# ===================================================================
# SHEET 5: RECONSTRUCTION_COMPLETENESS
# ===================================================================
RECONSTRUCTION_COMPLETENESS = [
    ["Layer", "Estimate", "Confidence", "Justification"],
    ("UI Layer (frontend behavior)", "90%", "HIGH", "Next.js SPA, JS bundles fully inspected. Dashboard, editor, and template UI behavior well understood. Only unverified: exact WASM editor UI behavior."),
    ("API Layer (endpoints)", "85%", "HIGH", "~25 endpoints discovered from bundle + official docs. Request/response shapes partially understood. Missing: exact error handling, webhook payloads, admin/internal endpoints."),
    ("Scoring Model (dimensions + logic)", "60%", "MEDIUM", "3 dimensions confirmed from official docs + bundle. 4th trend dimension inferred. Weights unknown. LLM vs ML model unknown. Have schema but not implementation."),
    ("Decision Engine (selection logic)", "55%", "MEDIUM", "Sort order confirmed. Task grouping confirmed. Clip metadata understood. But candidate generation, dedup strategy, and post-processing pipeline unclear."),
    ("Render Engine (output pipeline)", "45%", "LOW", "WASM engine known to exist. 4 output formats confirmed. Engine API endpoints found. But: WASM internal behavior not analyzed, server-side render pipeline unknown, encoding approach unknown."),
    ("Camera Engine (framing decisions)", "40%", "LOW", "GANYIQ port exists based on behavior observation. Actual Opus EMA parameters, threshold values, and transition logic unknown. Pattern likely correct, exact implementation unknown."),
    ("Infrastructure (servers, queue, DB)", "25%", "LOW", "Frontend stack well understood (Next.js + Express). Backend infrastructure (DB, queue, workers, GPUs) almost entirely unknown. Most claims are industry-standard guesses."),
    ("Business Model (pricing, retention)", "50%", "MEDIUM", "Pricing tiers understood (Free/Starter/Pro/Max/Enterprise). API pricing understood (credits). Virality Score premium feature confirmed. But unit economics, churn, growth metrics unknown."),
    
    ("OVERALL RECONSTRUCTION", "55%", "MEDIUM", "Frontend + API ≈ 85-90% complete. Core scoring engine ≈ 50-60% complete. Backend infrastructure ≈ 25% complete. Canvas/infrastructure layers are weakest. Average: ~55%"),
]

# ===================================================================
# SHEET 6: FALSE_CLAIMS — Claims that may be wrong
# ===================================================================
FALSE_CLAIMS = [
    ["Claim", "Risk of Being Wrong", "Why", "Impact if Wrong"],
    ("Estimated scoring weights (25-30%, 20-25%, etc)", "HIGH — Almost certainly wrong", "No evidence for any weight values. These were fabricated for the spreadsheet.", "Entire OPUS_SCORING_MODEL_V1 sheet would be misleading. Weights are the most important unknown."),
    ("Opus does NOT use ffmpeg", "MEDIUM — Correct for frontend, possibly wrong for server", "Frontend = confirmed. Server = unknown. Industry standard is ffmpeg for encoding.", "If Opus uses ffmpeg server-side, GANYIQ's ffmpeg pipeline is actually similar to Opus's server approach."),
    ("Audio event multipliers (1.0-1.5x)", "HIGH — GANYIQ values presented as Opus", "GANYIQ's interpretation, not extracted from Opus code.", "May mislead GANYIQ tuning priorities if Opus uses completely different weights."),
    ("NLE layer compositor (AVEditTimeline, etc)", "MEDIUM — May be partially correct but unverified", "Previous WASM analysis not re-verified. Symbols could be from embedded libraries.", "If Opus DOESN'T use NLE compositing, then the fundamental architecture gap claim is wrong."),
    ("PostgreSQL database", "HIGH — Zero evidence", "Pure assumption based on common stack.", "Low impact — DB choice doesn't affect product decisions."),
    ("Redis job queue", "HIGH — Zero evidence", "Pure assumption from async pattern.", "Low impact — queue technology doesn't affect product decisions."),
    ("4-dimension scoring (including Trend)", "MEDIUM — Trend may be incorrect", "Trend only in JS bundle, not in official docs. May be UI-only field or ClipAnything-only field.", "If Trend isn't a scoring dimension, our 4-dim model has 1 wrong dimension."),
    ("LLM-powered scoring (vs ML model)", "MEDIUM — Docs say 'AI' not 'LLM'", "Smaller ML models can do this without being LLMs.", "If it's a small ML model, GANYIQ could train one. If it's LLM, GANYIQ needs API access."),
]

# Write ALL sheets
write_sheet("CLAIM_REVIEW", CLAIM_REVIEW[0], CLAIM_REVIEW[1:])
write_sheet("EVIDENCE_CHAIN_REVIEW", EVIDENCE_CHAIN_REVIEW[0], EVIDENCE_CHAIN_REVIEW[1:])
write_sheet("DECISION_ENGINE_CONFIDENCE", DECISION_ENGINE_CONFIDENCE[0], DECISION_ENGINE_CONFIDENCE[1:])
write_sheet("UNKNOWNS_REGISTRY", UNKNOWNS_REGISTRY[0], UNKNOWNS_REGISTRY[1:])
write_sheet("RECONSTRUCTION_COMPLETENESS", RECONSTRUCTION_COMPLETENESS[0], RECONSTRUCTION_COMPLETENESS[1:])
write_sheet("FALSE_CLAIMS", FALSE_CLAIMS[0], FALSE_CLAIMS[1:])

wb.save(PATH)
print(f"\n✅ Red team review saved to: {PATH}")
print(f"Total sheets: {len(wb.sheetnames)}")
