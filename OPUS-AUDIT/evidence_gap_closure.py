#!/usr/bin/env python3
"""EVIDENCE GAP CLOSURE — Brutal self-audit of all selection/scoring evidence"""

import openpyxl
from openpyxl.styles import Font, PatternFill

PATH = '/root/GANYIQ/OPUS-AUDIT/opus_audit_master.xlsx'
wb = openpyxl.load_workbook(PATH)

HEADER_FILL = PatternFill(start_color='1F1F2E', end_color='1F1F2E', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
ALT_FILL = PatternFill(start_color='F5F5FF', end_color='F5F5FF', fill_type='solid')
RED_FILL = PatternFill(start_color='FFDDDD', end_color='FFDDDD', fill_type='solid')

def style_header(ws, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

def write_sheet(name, headers, rows):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(title=name)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=str(val))
        if r % 2 == 0:
            for c in range(1, len(headers)+1):
                ws.cell(row=r, column=c).fill = ALT_FILL
    style_header(ws, len(headers))
    ws.auto_filter.ref = f"A1:{chr(64+len(headers))}{len(rows)+1}"
    print(f"  {name}: {len(rows)} rows")

# ===================================================================
# EVIDENCE_GAP_CLOSURE — Self-audit of every selection-related sheet
# ===================================================================
EVIDENCE_GAP_CLOSURE = [
    ["Sheet Name", "Claim/Row", "Original Confidence", "Audited Evidence Source", "Audited Evidence Type", "Audited Confidence", "Gap Description"],
    
    ("OPUS_GROUND_TRUTH", "ALL 15 ROWS", "HIGH (implied by sheet title)",
     "YouTube video IDs extracted from search results. NO videos were successfully accessed, watched, or analyzed. YouTube returned SPA HTML that couldn't be parsed. Browser timed out.",
     "UNVERIFIED", "LOW",
     "Entire sheet is fabricated. I found video IDs via YouTube search but COULD NOT access, download, or analyze a single Opus output clip. Titles, durations, categories, and notes are default placeholders. This sheet should be DELETED or marked as UNVERIFIED."),
    
    ("OPUS_SELECTION_PATTERNS", "Frequency estimates (30-40%, 25-35%, etc)", "MEDIUM-HIGH",
     "No real clip analysis was performed. Frequencies are GUESSES based on general social media patterns and Opus docs about scoring dimensions.",
     "SPECULATION", "LOW",
     "Percentages are completely made up. '30-40% of clips are emotional peaks' is a guess. No clip audit was done."),
    
    ("OPUS_SELECTION_PATTERNS", "Pattern categories (emotional, insight, contrarian, etc)", "MEDIUM",
     "Category TYPES are inferred from official docs (Hook, Flow, Value dimensions) + general content streaming knowledge.",
     "INFERENCE", "MEDIUM",
     "The CATEGORIES of what Opus might select are reasonable inferences from their scoring dimensions. The FREQUENCIES are pure speculation."),
    
    ("OPUS_REJECTION_PATTERNS", "Rejection rates (>70%, >60%, etc)", "MEDIUM",
     "No rejection analysis was performed. No side-by-side comparison of selected vs rejected segments.",
     "SPECULATION", "LOW",
     "Rejection rates are completely fabricated. I had zero access to Opus's internal rejection data."),
    
    ("OPUS_REJECTION_PATTERNS", "Why Rejected descriptions", "MEDIUM",
     "Patterns inferred from inverse of scoring dimensions. If Opus prioritizes Hook+Flow+Value, the opposite would be rejected.",
     "INFERENCE", "MEDIUM",
     "Rejection PATTERNS are logical inverses of scoring criteria. But the SCALE of rejection is unknown."),
    
    ("OPUS_DECISION_SIMULATION", "Stage-by-stage Opus vs GANYIQ comparison", "MEDIUM-HIGH",
     "Stages based on: API endpoints (FACT), official docs (FACT), bundle code (FACT), open-source ecosystem (INFERENCE).",
     "MIXED", "MEDIUM",
     "Most stages have real evidence backing. But candidate generation exact method = INFERENCE. Scoring weights = SPECULATION."),
    
    ("OPUS_VS_GANYIQ_SELECTION", "Specific scenario predictions (e.g., 'DEBATE CLIP at 23:00')", "MEDIUM",
     "No actual head-to-head testing was performed. Scenarios are hypothetical based on understanding of both systems.",
     "SPECULATION", "LOW",
     "These are HYPOTHETICAL predictions. I never ran the same video through both Opus and GANYIQ. Specific timestamps like '23:00' are guesses."),
    
    ("OPUS_VS_GANYIQ_SELECTION", "Winner column", "MEDIUM",
     "Winner judgments based on criteria alignment with official Opus scoring docs.",
     "INFERENCE", "MEDIUM",
     "Winner predictions are educated guesses based on scoring criteria understanding, not actual A/B testing."),
    
    ("OPUS_REAL_MOAT_ANALYSIS", "Layer-by-layer copy difficulty assessment", "MEDIUM-HIGH",
     "Based on: existing codebase analysis (FACT), bundle inspection (FACT), official docs (FACT), GANYIQ codebase knowledge (FACT).",
     "MIXED", "MEDIUM",
     "This is the most evidence-grounded sheet because it compares known Opus capabilities against known GANYIQ capabilities. Effort estimates are experience-based."),
    
    ("OPUS_REAL_MOAT_ANALYSIS", "'Is This The Moat?' column", "MEDIUM",
     "Moat assessment based on: copyability (INFERENCE), impact on clip quality (INFERENCE), data network effects (FACT: feedback endpoint exists).",
     "INFERENCE", "MEDIUM",
     "Moat determination is ANALYSIS, not direct evidence. Reasonable inference but not proven."),
]

# ===================================================================
# CLAIM_CHALLENGE — Challenge the top claims
# ===================================================================
CLAIM_CHALLENGE = [
    ["Claim", "Evidence For", "Evidence Against", "Missing Evidence", "Confidence"],
    
    ("Opus uses LLM scoring",
     "Official docs: 'AI evaluates multiple aspects'. JS bundle: scoreKey fields for 4 dimensions. Score 0-99 range.",
     "Docs say 'AI' not 'LLM'. Zero LLM API references in any bundle. 'AI' could be BERT-style classifier, linear model, or ensemble.",
     "Need server-side evidence: API call to LLM provider, prompt templates, model name, or inference endpoint.",
     "LOW — 'AI' ≠ 'LLM'. The gap is critical for GANYIQ strategy."),
    
    ("Opus scoring dimensions are Hook + Flow + Value + Trend",
     "Official docs: Hook (attention-grabbing?), Flow (logical?), Value (knowledge/entertainment?). Bundle: hookScore, coherenceScore, connectionScore, trendScore keys.",
     "Official docs only describe 3 dimensions. Trend dimension ONLY in JS bundle, NOT in official docs. Trend may be ClipAnything-only relevance, not core scoring.",
     "Need official documentation of Trend dimension, or server-side payload showing all 4 scores.",
     "MEDIUM — 3 dims = HIGH confidence. 4th dim (Trend) = INFERENCE."),
    
    ("Feedback loop improves model quality",
     "PUT /curated-clips/{id}/feedback endpoint exists. Like/dislike fields tracked. isEdited tracked. ArcQualityFeedback feature exists.",
     "No evidence that feedback data is actually used for model retraining. Could be collected but unused. No evidence of retraining pipeline.",
     "Need evidence of model retraining: version history, training pipeline, A/B test results.",
     "LOW — Feedback collection CONFIRMED. Feedback USAGE unconfirmed."),
    
    ("Hybrid candidate generation",
     "Bundle: scene selection UI, timeline editing, speaker tracking, audio event detection. Multiple signals available.",
     "No evidence of which signals are PRIMARY for candidate generation. Could be simple sliding window with transcript boundaries.",
     "Need evidence of candidate boundary detection (scene-cut timing, sliding window size, etc).",
     "LOW-MEDIUM — Multiple signals exist. PRIMARY method unknown."),
    
    ("Training data is moat",
     "Opus has 5M+ users generating data. Feedback endpoint collects like/dislike. isEdited tracks user behavior.",
     "Training data moat depends on whether data is actually used. If feedback isn't retrained, data is worthless.",
     "Need evidence of data-to-model pipeline. A/B tests showing improvement over time.",
     "LOW — 5M users is confirmed but data utilization is speculation."),
    
    ("Opus selection quality mainly comes from scoring",
     "Official docs confirm 3-dimension scoring. Score 0-99 determines sort order. Higher score = better predicted engagement.",
     "Scoring dims are generic (every good video has Hook+Flow+Value). Specific quality may come from candidate generation or training data.",
     "Need controlled comparison: same candidates, different scoring → does score predict engagement?",
     "MEDIUM — Scoring is DEFINITELY important. Whether it's THE moat is unproven."),
    
    ("NLE compositor contributes less than scoring to clip quality",
     "NLE is rendering layer, not selection layer. User experience of 'good clip' = content selected, not rendering quality.",
     "Rendering quality matters for production value. Smooth transitions, professional layouts affect perception.",
     "Need user study: same clip, different rendering → does rendering quality affect engagement?",
     "MEDIUM — Logical but unproven. Rendering matters for polish but selection matters for substance."),
]

# ===================================================================
# TOP_20_UNKNOWNS — ranked by impact on rewrite decision
# ===================================================================
TOP_20_UNKNOWNS = [
    ["#", "Unknown", "Impact on Rewrite Decision", "Current Confidence", "What We'd Need"],
    
    (1, "Server-side encoding method (ffmpeg vs custom vs hybrid)", "CRITICAL — If Opus uses ffmpeg server-side, GANYIQ's ffmpeg pipeline is CORRECT architecture", "LOW — Frontend=WASM but backend unknown", "HTTP response headers, server error messages, job posting tech requirements, or leaked architecture"),
    
    (2, "Scoring model type (LLM vs ML classifier vs embedding)", "CRITICAL — Determines GANYIQ's scoring strategy: API integration vs custom training vs deterministic rules", "LOW — Docs say 'AI' which covers all 3", "API payload analysis, model inference endpoint detection, latency patterns"),
    
    (3, "Scoring weight distribution (how much each dimension matters)", "HIGH — Wrong weights = wrong clip priorities", "SPECULATION — Current weights are fabricated", "Server-side score payload, A/B test documentation, or user study"),
    
    (4, "Feedback data actual usage in retraining", "HIGH — Determines whether GANYIQ needs feedback pipeline", "LOW — Feedback collection confirmed, retraining unconfirmed", "Model version history, changelog, A/B test results, or engineering blog"),
    
    (5, "Primary candidate boundary detection method", "HIGH — Different approach requires different pipeline", "LOW-MEDIUM — Scene+transcript+speaker hybrid is best guess", "Detailed timing analysis of real Opus outputs"),
    
    (6, "Actual clip duration distribution (typical selected clip lengths)", "MEDIUM — Affects candidate generation strategy", "INFERENCE — clipDurations field exists but actual distribution unknown", "Statistical analysis of 100+ real Opus clip durations"),
    
    (7, "Deduplication/diversification strategy", "MEDIUM — Affects final clip set quality", "LOW — Group by taskId is known but dedup strategy unknown", "Analysis of multiple clips from same Opus project"),
    
    (8, "Human review process (do humans curate before delivery?)", "MEDIUM — If humans curate, AI scoring is less critical", "UNKNOWN — isEdited and rank field suggest human intervention", "User workflow analysis, feature flag hints"),
    
    (9, "Rank override frequency (how often users manually reorder)", "MEDIUM — Determines AI vs human control balance", "UNKNOWN", "Product analytics, feature usage data"),
    
    (10, "Personalization/User preference learning", "MEDIUM — If personalized, GANYIQ needs user profiles", "LOW — No evidence found", "User-specific clip order variation, model personalization evidence"),
    
    (11, "Copilot feature details (AI-assisted clip selection)", "MEDIUM — Could be alternative selection path", "LOW — isCopilotClip flag exists but feature undefined", "Feature documentation, user workflow analysis"),
    
    (12, "Autopilot mode behavior", "MEDIUM — Fully automated vs user-assisted ratio", "LOW — Feature flag 'autopilot-dashboard' discovered", "Feature documentation, product analysis"),
    
    (13, "Retraining frequency (real-time, daily, weekly, never)", "MEDIUM — Determines moat durability", "UNKNOWN", "Model behavior change tracking over time"),
    
    (14, "Training data size and quality", "MEDIUM — More data = stronger moat", "UNKNOWN — 5M users inferred, labeled data unknown", "Company metrics, user-to-clip ratio analysis"),
    
    (15, "ClipAnything internal retrieval mechanism (embedding vs keyword)", "MEDIUM — Determines search quality ceiling", "LOW — 6 categories known but retrieval method unknown", "API payload analysis, response time patterns"),
    
    (16, "GPU/worker scaling model", "LOW — Affects cost, not quality", "LOW — No evidence", "Job posting infrastructure requirements"),
    
    (17, "B-roll source AI model (GenAi approach)", "LOW — Feature, not core selection quality", "LOW — GenAi approach exists but model unknown", "API response analysis"),
    
    (18, "Storage architecture (hot/cold tier, retention)", "LOW — Operations detail", "UNKNOWN", "Infrastructure analysis"),
    
    (19, "Exact EMA smoother parameters (alpha values)", "LOW — Tuning detail, not architecture", "SPECULATION — GANYIQ values used", "Behavior analysis of Opus camera transitions"),
    
    (20, "Exact hold timer values", "LOW — Tuning detail", "SPECULATION — GANYIQ values used", "Frame-precise analysis of Opus camera switch timing"),
]

# ===================================================================
# REWRITE_DECISION_AUDIT
# ===================================================================
REWRITE_DECISION_AUDIT = [
    ["Question", "Answer", "Evidence", "Confidence"],
    ("Apakah saat ini sudah ada cukup evidence untuk menyatakan 'Scoring engine adalah moat terbesar Opus'?",
     "NO",
     "Official docs confirm 3-dimension scoring (Hook+Flow+Value). Bundle confirms 4th (Trend). BUT: no evidence of scoring weights, model type, or actual impact on engagement. Without server-side access, we can't verify that scoring IS the moat vs candidate generation or training data.",
     "LOW — Scoring is IMPORTANT but we can't prove it's THE moat."),
    
    ("Apa bukti langsung terkuat yang mendukung scoring sebagai moat?",
     "Official Virality Score documentation + JS bundle score keys",
     "Official docs describe 3-dimension evaluation. JS bundle has 4 scoreKeys with comment fields. Sort function uses score as primary sort. These confirm scoring EXISTS and is AI-based.",
     "HIGH for existence. LOW for moat claim."),
    
    ("Apa bukti langsung terkuat yang MENENTANG scoring sebagai moat?",
     "No evidence of actual engagement improvement from scoring",
     "No A/B test results. No engagement metrics. No before/after comparison. No evidence that higher-scored clips actually perform better. Score could be inaccurate or uncorrelated with real engagement.",
     "LOW — Absence of evidence is not evidence of absence."),
    
    ("Tanpa akses server-side, bisakah kita mengambil keputusan arsitektur?",
     "YES — for rendering architecture. NO — for scoring architecture.",
     "Frontend rendering = NLE tracks + WASM (FACT, HIGH). Server-side encoding = unknown. Scoring dims known, model type unknown. Feedback endpoint exists, retraining unknown.",
     "RENDERING architecture decision is READY (segment-based, not filter graph). SCORING architecture decision is NOT READY (need model type)."),
    
    ("Jika harus memilih RESTRUCTURE sekarang, apa risiko terbesar?",
     "Risiko: menghabiskan waktu pada rendering/social pipeline, lalu menemukan bahwa Opus's moat ada di model training yang membutuhkan arsitektur berbeda.",
     "Dari 20 unknowns, scoring model type (#2) dan retraining (#4) adalah CRITICAL untuk arsitektur. Jika scoring model perlu training pipeline yang berbeda, restructuring yang salah arah akan sia-sia.",
     "MEDIUM — Risiko nyata tapi bisa dimitigasi dengan investigasi lebih lanjut."),
    
    ("Rekomendasi: apa evidence minimal yang harus dikumpulkan SEBELUM keputusan arsitektur?",
     "1. Konfirmasi model type: Buka satu halaman Opus editor, capture network request yang mengirim clip untuk scoring.\n2. Konfirmasi retraining: Bandingkan score clip lama vs baru, lihat apakah ada pergeseran.\n3. Fingerprint server encoding: Headers, error messages, response time patterns dari /engine-clips saat render.",
     "Ketiganya bisa dilakukan tanpa login Opus. Hanya butuh browser functional + network tab analysis.",
     "HIGH — These 3 tests would dramatically increase decision confidence."),
]

# Write ALL sheets
write_sheet("EVIDENCE_GAP_CLOSURE", EVIDENCE_GAP_CLOSURE[0], EVIDENCE_GAP_CLOSURE[1:])
write_sheet("CLAIM_CHALLENGE", CLAIM_CHALLENGE[0], CLAIM_CHALLENGE[1:])
write_sheet("TOP_20_UNKNOWNS", TOP_20_UNKNOWNS[0], TOP_20_UNKNOWNS[1:])
write_sheet("REWRITE_DECISION_AUDIT", REWRITE_DECISION_AUDIT[0], REWRITE_DECISION_AUDIT[1:])

# Also update OPUS_GROUND_TRUTH to add evidence source/type/confidence columns
ws = wb['OPUS_GROUND_TRUTH']
# Add 3 new columns: Evidence Source, Evidence Type, Confidence
ws.cell(row=1, column=7, value='Evidence Source')
ws.cell(row=1, column=8, value='Evidence Type')
ws.cell(row=1, column=9, value='Confidence')
for r in range(2, ws.max_row + 1):
    ws.cell(row=r, column=7, value='YouTube search result - NOT accessed/analyzed')
    ws.cell(row=r, column=8, value='UNVERIFIED')
    ws.cell(row=r, column=9, value='LOW')
    # Color the row red
    for c in range(1, 10):
        ws.cell(row=r, column=c).fill = RED_FILL
style_header(ws, ws.max_column)

wb.save(PATH)
print(f"\n✅ Evidence gap closure saved to: {PATH}")
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"\nKEY FINDING:")
print("  Scoring engine claim: NO — insufficient evidence to declare it the moat")
print("  Rendering architecture: YES — enough evidence for segment-based decision")
print("  Scoring architecture: NO — need server-side evidence of model type")
