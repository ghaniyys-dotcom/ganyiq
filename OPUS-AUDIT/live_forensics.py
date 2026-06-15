#!/usr/bin/env python3
"""LIVE FORENSICS — Evidence collected from real-time probes on June 22 2026"""

import openpyxl
from openpyxl.styles import Font, PatternFill

PATH = '/root/GANYIQ/OPUS-AUDIT/opus_audit_master.xlsx'
wb = openpyxl.load_workbook(PATH)

HEADER_FILL = PatternFill(start_color='1F1F2E', end_color='1F1F2E', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
ALT_FILL = PatternFill(start_color='F5F5FF', end_color='F5F5FF', fill_type='solid')

def style_header(ws, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

def write_sheet(name, headers, rows):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(title=name)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=str(val))
        if r % 2 == 0:
            for c in range(1, len(headers)+1):
                ws.cell(row=r, column=c).fill = ALT_FILL
    style_header(ws, len(headers))
    ws.auto_filter.ref = f"A1:{chr(64+len(headers))}{len(rows)+1}"
    print(f"  {name}: {len(rows)} rows")

# ===================================================================
# LIVE_NETWORK_FORENSICS
# ===================================================================
LIVE_NETWORK = [
    ["Endpoint/URL", "Method", "Status", "Evidence Type", "Details", "Confidence"],
    ("clip.opus.pro/", "GET", "200 (5.5KB HTML)", "FACT", "Next.js Pages Router SPA. Build ID: j5ggUop8Nh8jeyeoNqQYM. 10M+ users meta description.", "HIGH"),
    ("api.opus.pro/api/health", "GET", "200", "FACT", "API health check endpoint responds with 200. Express-based backend confirmed.", "HIGH"),
    ("public.cdn.opus.pro/", "GET", "403", "FACT", "CDN domain returns 403 without auth. Cloudflare-protected.", "HIGH"),
    ("public.gcs.opus.pro/", "GET", "403", "FACT", "Google Cloud Storage bucket. 4KB 403 response (GCS-style).", "HIGH"),
    ("eq.opus.pro/", "GET", "403", "FACT", "Subdomain returns 403. Likely Engine Queue service, auth-protected.", "HIGH"),
    ("agent.opus.pro/", "GET", "200 (16KB HTML)", "FACT", "Agent Opus — separate Next.js App Router site. Build SHA: e9984ae. Separate CDN: prod-ao-ext.cdn.opus.pro.", "HIGH"),
    ("auth.opus.pro/", "GET", "Timeout", "INFERENCE", "Auth service did not respond. Likely internal service, not publicly accessible.", "MEDIUM"),
    ("stg-eq.opus.pro/", "GET", "Not tested", "UNVERIFIED", "Staging Engine Queue subdomain found in previous analysis.", "LOW"),
    ("stg-agent.eng.opus.pro/", "GET", "Not tested", "UNVERIFIED", "Staging Agent Engineering subdomain found previously. 'eng' = engineering namespace.", "LOW"),
    ("help.opus.pro/docs/article/virality-score", "GET", "200", "FACT", "Official virality score documentation retrieved. 3 dimensions confirmed. Mintlify-hosted.", "HIGH"),
    ("_next/static/chunks/9201-d815e8c0de87d576.js", "GET", "200 (40KB)", "FACT", "New JS chunk not previously analyzed. Contains Google Auth + Intercom integration.", "HIGH"),
    ("accounts.google.com/gsi/client", "GET", "200", "FACT", "Google OAuth client loaded on clip.opus.pro.", "HIGH"),
    ("consent.cookiebot.com/uc.js", "GET", "200", "FACT", "CookieBot GDPR consent loaded.", "HIGH"),
    ("featureassets.org", "GET", "Likely 200", "INFERENCE", "A/B testing feature assets domain, preconnected in page source.", "MEDIUM"),
    ("o4504370498437120.ingest.sentry.io", "GET", "Connected", "FACT", "Sentry error tracking with dsn: 5bc42aea68214138b9671bfcb10bcb0d@o4504370498437120", "HIGH"),
    ("public.cdn.opus.pro/clip-web/wasm/AVEditorEngine-*.wasm.gz", "GET", "200 (5.96MB)", "FACT", "AVEditorEngine WASM confirmed at CDN path. Build date June 4 2026.", "HIGH"),
]

# ===================================================================
# LIVE_EDITOR_FORENSICS
# ===================================================================
LIVE_EDITOR = [
    ["Field/Component", "Evidence", "Source", "Type", "Confidence"],
    ["editingScript?.tracks", "editingScript has tracks property. ?include=editingScript param fetches it.", "/tmp/app.js — exportable-clips API", "FACT", "HIGH"],
    ["renderAsVideoPreview", "Preview render output URL field in clip data model", "/tmp/app.js", "FACT", "HIGH"],
    ["renderAsVideoFile", "Standard export URL field", "/tmp/app.js", "FACT", "HIGH"],
    ["renderAsVideoFile4K", "4K export URL field", "/tmp/app.js", "FACT", "HIGH"],
    ["renderAsAdobeXml", "Adobe Premiere Pro XML export", "/tmp/app.js", "FACT", "HIGH"],
    ["uriForPreview", "Preview URI", "/tmp/app.js", "FACT", "HIGH"],
    ["uriForExport", "Export URI", "/tmp/app.js", "FACT", "HIGH"],
    ["uriForExport4K", "4K export URI", "/tmp/app.js", "FACT", "HIGH"],
    ["uriForAdobePr", "Adobe Premiere URI", "/tmp/app.js", "FACT", "HIGH"],
    ["captions", "enableCaption, captionStyleId, enableCaptionAnimation, enableHighlight, enableEmoji", "/tmp/app.js", "FACT", "HIGH"],
    ["timeline", "selectedTimeline, editedTimeline, timelineIn fields in brolls API. Positions in ms.", "/tmp/app.js", "FACT", "HIGH"],
    ["broll", "B-roll system: Stock (Pixels/Storyblocks/Shutterstock) + GenAi (AI). Auto/Select/Prompt.", "/tmp/app.js", "FACT", "HIGH"],
    ["manual reframing", "Manual reframing feature: help article URL, jobType='reframe'. User can adjust AI crops.", "/tmp/app.js", "FACT", "HIGH"],
    ["render elements", "'original-video-render', 'auto-tracking-video-render', 'auto-tracking-segment-render', 'add-section'", "/tmp/app.js CSS class names", "FACT", "MEDIUM"],
]

# ===================================================================
# LIVE_SCORING_FORENSICS
# ===================================================================
LIVE_SCORING = [
    ["Evidence", "Source", "Classification", "Confidence", "Notes"],
    ("Virality Score range: 0-99", "OFFICIAL DOCS: help.opus.pro virality-score page", "FACT", "HIGH", "Directly stated in official help docs."),
    ("3 scoring dimensions: Hook + Flow + Value", "OFFICIAL DOCS: Hook: grab attention. Flow: logical flow. Value: emotional connection.", "FACT", "HIGH", "ALL 3 confirmed in official documentation."),
    ("Value dimension = emotional connection", "OFFICIAL DOCS: 'Does the video offer value, resonate emotionally, and create a personal connection with the audience?'", "FACT", "HIGH", "PREVIOUSLY MISINTERPRETED. Value is about EMOTIONAL CONNECTION, not 'actionable advice'. This changes GANYIQ strategy."),
    ("Hook description", "OFFICIAL DOCS: 'Does the introduction grab attention and directly relate to the main topic of the video?'", "FACT", "HIGH", "Hook is about attention + relevance to topic."),
    ("Flow description", "OFFICIAL DOCS: 'Does the video flow logically from one part to the next, with a satisfying conclusion?'", "FACT", "HIGH", "Flow is about narrative logic and payoff."),
    ("AI evaluates", "OFFICIAL DOCS: 'OpusClip's AI evaluates multiple aspects of your video'", "FACT", "HIGH", "Uses 'AI' not 'LLM'. Important distinction."),
    ("Default sort by score descending", "OFFICIAL DOCS: 'By default, your clips will be sorted by Virality Score, from highest to lowest.'", "FACT", "HIGH", "Default sort order confirmed."),
    ("Pro/Starter only feature", "OFFICIAL DOCS: 'available exclusively to Pro and Starter plan users'", "FACT", "HIGH", "Free tier users don't see scores."),
    ("Prompt relevance modifier", "OFFICIAL DOCS: 'evaluates whether the clip is relevant to your prompt, when using ClipAnything model'", "FACT", "HIGH", "Additional modifier in ClipAnything mode."),
    ("Score key fields in JS bundle", "hookScore, coherenceScore, connectionScore, trendScore with commentKey fields", "FACT", "HIGH", "4th dimension (trendScore) in bundle but NOT in official docs."),
    ("judgeResult object", "judgeResult: {isCopilotClip, relevanceScore, trendTopic, score}", "FACT", "HIGH", "Contains additional scoring metadata."),
    ("Feedback endpoint: like/dislike", "PUT /curated-clips/{id}.{archId}/feedback", "FACT", "HIGH", "User feedback collection confirmed."),
    ("10M+ users", "Meta description in page source: '10M+ creators'", "FACT", "HIGH", "Directly from Opus's own meta description."),
    ("3 dimensions ONLY (not 4)", "Official docs list ONLY Hook+Flow+Value. Trend is NOT in official docs.", "FACT", "HIGH", "PREVIOUSLY OVERCLAIMED. I said '4 dimensions confirmed' — WRONG. Only 3 are confirmed in official docs. Trend is bundle-only."),
]

# ===================================================================
# LIVE_CANDIDATE_FORENSICS
# ===================================================================
LIVE_CANDIDATE = [
    ["Evidence", "Source", "Classification", "Confidence", "Notes"],
    ("curated-clips API", "GET /curated-clips/feedback?q=findByProjectId&projectId= — returns clip list with scores", "/tmp/app.js", "FACT", "Clips are curated before export. Separate curation stage."),
    ("exportable-clips API", "GET /exportable-clips?projectId= — separate endpoint for export-ready clips", "/tmp/app.js", "FACT", "Curation → Export is a 2-stage pipeline."),
    ("engine-clips API", "PUT/POST/GET /engine-clips/{id} — rendering stage endpoint", "/tmp/dashboard.js", "FACT", "Engine stage handles rendering. Separate from curation."),
    ("clipDurations array", "curationPref.clipDurations — user sets preferred clip lengths", "API schema", "FACT", "User-defined durations influence candidate filtering."),
    ("topicKeywords array", "curationPref.topicKeywords — keywords filter candidates", "API schema", "FACT", "Topics must be present in transcript."),
    ("genre field", "genre in clip data model. curation_api_config with ready[] and beta[] genres.", "/tmp/app.js", "FACT", "Genre-specific curation models."),
    ("b-roll timeline integration", "selectedTimeline + editedTimeline fields in brolls API. timelineIn precision.", "/tmp/app.js", "FACT", "Candidates have timeline positions for b-roll insertion."),
    ("Scene/timeline arrangement UI", "User can 'Arrange Scenes in Timeline', drag segment boundaries. Visual timeline editor.", "/tmp/app.js string literals", "FACT", "Visual timeline editing of candidates."),
    ("Candidate generation model UNKNOWN", "No evidence of primary chunking method (sliding window, speaker-turn, scene-detection, etc.)", "N/A", "SPECULATION", "Still unknown how initial candidates are generated."),
]

# ===================================================================
# LIVE_BACKEND_FORENSICS
# ===================================================================
LIVE_BACKEND = [
    ["Evidence", "Source", "Classification", "Confidence", "Notes"],
    ("ffmpeg references in frontend: ZERO", "0 hits across all bundles", "/tmp/app.js, dashboard.js, common-ui.js, main.js, chunk9201.js", "FACT", "Frontend has zero ffmpeg references."),
    ("WASM engine on CDN: AVEditorEngine", "public.cdn.opus.pro/clip-web/wasm/AVEditorEngine-20260604-09125d5a.wasm.gz (5.96MB)", "CDN probe", "FACT", "Frontend uses WASM for composition."),
    ("Editing script with tracks", "editingScript?.tracks — NLE-style track-based composition", "/tmp/app.js", "FACT", "Track-based editing script format."),
    ("Engine API endpoints (5 endpoints)", "PUT/{id}, POST/render, POST/save-and-render, POST/refine, GET/{id}", "/tmp/dashboard.js", "FACT", "Server-side async render pipeline."),
    ("eq.opus.pro subdomain", "eq.opus.pro returns 403. Likely Engine Queue service.", "DNS/subdomain probe", "INFERENCE", "Separate queue service for render jobs."),
    ("Queue features: fast_queue, no_credit_required", "Priority queue strings in bundle", "/tmp/app.js", "FACT", "Multiple queue priority tiers."),
    ("4 render output formats", "renderAsVideoPreview, VideoFile, VideoFile4K, AdobeXml", "/tmp/app.js", "FACT", "Multi-format output."),
    ("Render element names", "'original-video-render', 'auto-tracking-video-render', 'auto-tracking-segment-render', 'add-section'", "/tmp/app.js", "FACT", "Multiple render paths exist."),
    ("Clip export tracking", "clip-export-records POST/PUT/GET. check-export-operation.", "/tmp/app.js", "FACT", "Export monitoring system."),
    ("Server encoding method: UNKNOWN", "No evidence of ffmpeg, custom encoder, or cloud API on server-side.", "N/A", "SPECULATION", "Most critical unknown. Encoding method determines architecture decision."),
    ("GPU/NVENC/CUDA: ZERO evidence", "No references in any probe", "All sources", "SPECULATION", "No evidence for or against GPU encoding."),
]

# Write ALL sheets
write_sheet("LIVE_NETWORK_FORENSICS", LIVE_NETWORK[0], LIVE_NETWORK[1:])
write_sheet("LIVE_EDITOR_FORENSICS", LIVE_EDITOR[0], LIVE_EDITOR[1:])
write_sheet("LIVE_SCORING_FORENSICS", LIVE_SCORING[0], LIVE_SCORING[1:])
write_sheet("LIVE_CANDIDATE_FORENSICS", LIVE_CANDIDATE[0], LIVE_CANDIDATE[1:])
write_sheet("LIVE_BACKEND_FORENSICS", LIVE_BACKEND[0], LIVE_BACKEND[1:])

wb.save(PATH)
print(f"\n✅ Live forensics saved to: {PATH}")
print(f"Total sheets: {len(wb.sheetnames)}")

print("\n=== KEY CORRECTIONS FROM LIVE FORENSICS ===")
print("1. Value dimension = 'emotional connection' (NOT 'actionable advice')")
print("   OLD: 'valuable knowledge, actionable advice, or entertaining content'")
print("   NEW: 'offer value, resonate emotionally, create personal connection with audience'")
print()
print("2. Only 3 dimensions confirmed (NOT 4)")
print("   OLD: '4-dimension scoring: Hook + Flow + Value + Trend'")
print("   NEW: '3-dimension scoring: Hook + Flow + Value. Trend is bundle-only.'")
print()
print("3. Docs say 'AI' not 'LLM'")
print("   OLD: assumed LLM")
print("   NEW: 'AI evaluates' — could be smaller ML model")
