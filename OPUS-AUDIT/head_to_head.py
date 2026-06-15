#!/usr/bin/env python3
"""OPUS VS GANYIQ — Head to Head Clip Selection Forensics"""

import openpyxl
from openpyxl.styles import Font, PatternFill

PATH = '/root/GANYIQ/OPUS-AUDIT/opus_audit_master.xlsx'
wb = openpyxl.load_workbook(PATH)

HEADER_FILL = PatternFill(start_color='1F1F2E', end_color='1F1F2E', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
ALT_FILL = PatternFill(start_color='F5F5FF', end_color='F5F5FF', fill_type='solid')

def style_header(ws, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

def write_sheet(name, headers, rows):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(title=name)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=str(val))
        if r % 2 == 0:
            for c in range(1, len(headers)+1):
                ws.cell(row=r, column=c).fill = ALT_FILL
    style_header(ws, len(headers))
    ws.auto_filter.ref = f"A1:{chr(64+len(headers))}{len(rows)+1}"
    print(f"  {name}: {len(rows)} rows")

# ===================================================================
# OPUS_SELECTION_PATTERNS — What Opus selects
# ===================================================================
OPUS_SELECTION_PATTERNS = [
    ["Pattern", "Frequency Estimate", "Why Selected", "Hook Type", "Evidence", "GANYIQ Parallel"],
    ("Emotional Peak (laughter, gasp, strong reaction)", "HIGH — 30-40% of clips", "Strong emotional reactions = highest engagement. Laughter = 1.5x multiplier in ReactionScheduler.", "Surprise/Shock/Humor", "Official: PeakMomentDetector. Bundle: REACTION_WEIGHTS laughter=1.5, gasp=1.2, emotion_peak=1.4", "Port matched (GANYIQ has Ported ReactionScheduler) but trigger thresholds may differ"),
    ("High-value insight/actionable advice", "HIGH — 25-35% of clips", "Value dimension directly measures 'valuable knowledge, actionable advice, or entertaining content'.", "Authority/Lesson", "Official: Value dimension in Virality Score. Bundle: connectionScore.", "GANYIQ engagement scoring partially covers this but without LLM evaluation of actual value density"),
    ("Contrarian/Conflict opinion", "MEDIUM-HIGH — 15-25% of clips", "Contrarian opinions drive comments, shares, and debate. Multiple speakers = higher engagement.", "Contrarian/Debate", "Bundle: SpeakerActivityTracker. Trend dimension evaluates current relevance.", "Not evaluated by GANYIQ's deterministic scoring"),
    ("Narrative hook (story opening)", "MEDIUM — 10-20% of clips", "Hook dimension directly evaluates 'Does the introduction grab attention?'", "Story/Curiosity", "Official: Hook dimension in Virality Score.", "GANYIQ has hook_position scoring but no LLM evaluation of hook quality"),
    ("Trending topic alignment", "MEDIUM — 10-15% of clips", "Trend dimension evaluates relevance to current topics. trendTopic field in judgeResult.", "Trend/Prediction", "Bundle: trendScore + trendTopic in judgeResult. Official: ClipAnything 'viral topics' search.", "GANYIQ has NO trend evaluation at all"),
    ("Speaker switch / interruption", "MEDIUM — 10-15% of clips", "Speaker switches signal dynamic conversation. Each switch is a potential clip boundary.", "Conflict/Debate", "Bundle: SpeakerActivityTracker. turnDetected triggers peak detection.", "GANYIQ tracks speaker changes but doesn't use as selection signal"),
    ("Specific number/statistic reveal", "MEDIUM — 8-12% of clips", "Numbers are concrete, quotable, and highly shareable on social media.", "Data/Money", "Bundle: No direct evidence. Inferred from Value dimension + general social media patterns.", "Not explicitly tracked by GANYIQ"),
    ("Controversial statement", "MEDIUM — 8-12% of clips", "Controversy drives engagement. Trend dimension may capture controversy as 'trending'.", "Controversy/Shock", "Bundle: trendScore dimension. PeakMomentDetector escalation levels.", "Not evaluated by GANYIQ"),
    ("Prediction/Future claim", "LOW-MEDIUM — 5-10% of clips", "Predictions create curiosity about outcome, driving future engagement.", "Prediction", "Bundle: trendScore (future-oriented content). Inferred from social media patterns.", "Not evaluated by GANYIQ"),
    ("Vulnerability/Failure story", "LOW-MEDIUM — 5-10% of clips", "Failure stories are unexpectedly popular due to relatability and learning value.", "Mistake/Story", "Bundle: connectionScore (emotional connection).", "Partially covered by engagement scoring"),
]

# ===================================================================
# OPUS_REJECTION_PATTERNS — What Opus rejects
# ===================================================================
OPUS_REJECTION_PATTERNS = [
    ["Pattern", "Estimated Rejection Rate", "Why Rejected", "Evidence", "GANYIQ Difference"],
    ("Flat delivery, monotone, no vocal energy", "HIGH — >70% rejected", "No emotional peaks = no hook = no engagement. AI evaluates vocal energy and emotional modulation.", "Bundle: PeakMomentDetector requires score >50 to escalate. Flat delivery = score <50.", "GANYIQ doesn't measure vocal energy"),
    ("Repetitive content (same point rephrased)", "HIGH — >60% rejected", "No new value per second. Value dimension penalizes low information density.", "Bundle: Flow dimension 'logical flow'. Repeating = no flow.", "GANYIQ doesn't detect content repetition"),
    ("Meandering/No clear point", "HIGH — >60% rejected", "Flow dimension requires 'logical flow with satisfying conclusion'. No flow = low score.", "Official: Flow dimension evaluates logical structure.", "GANYIQ coherence measure is mathematical, not semantic"),
    ("Low information density", "MEDIUM-HIGH — 40-60% rejected", "Value dimension measures per-second value. Filler content dilutes value.", "Official: Value dimension 'valuable knowledge per second'.", "GANYIQ doesn't measure information density"),
    ("Weak/No hook in opening seconds", "MEDIUM-HIGH — 40-60% rejected", "Hook dimension directly evaluates first few seconds. No hook = immediate rejection.", "Official: Hook dimension 'Does the introduction grab attention?'", "GANYIQ has hook_position factor but no hook quality evaluation"),
    ("Context-setting (background, introductions)", "MEDIUM — 30-50% rejected", "Setup rarely has hook, value, or emotional content. AI skips to the 'good part'.", "Bundle: Most selected clips start mid-conversation, not at intro.", "GANYIQ may select these if energy score is high enough"),
    ("Off-topic tangents", "MEDIUM — 30-50% rejected", "Value + Hook dimensions penalize content unrelated to main topic.", "Bundle: value dimension + ClipAnything prompt relevance.", "GANYIQ doesn't evaluate topic relevance"),
    ("Audio event absence (no laughter, no emphasis)", "MEDIUM — 25-40% rejected", "Audio events trigger peak detection and layout escalation. No events = no escalation.", "Bundle: REACTION_WEIGHTS + PeakMomentDetector thresholds.", "GANYIQ has same detection but triggers are GANYIQ-tuned"),
    ("Single speaker, no interaction", "LOW-MEDIUM — 20-30% rejected", "Single speaker = less dynamic = lower engagement potential.", "Bundle: SpeakerActivityTracker prefers multiple active speakers.", "GANYIQ treats single speakers equally"),
    ("Low energy/debate quality", "LOW-MEDIUM — 15-25% rejected", "Low energy conversation doesn't trigger audio events or peak detection.", "Bundle: PeakMomentDetector minimum threshold of 50 for escalation.", "GANYIQ energy calculation may catch some but is rule-based"),
]

# ===================================================================
# OPUS_DECISION_SIMULATION — Reverse-engineered decision flow
# ===================================================================
OPUS_DECISION_SIMULATION = [
    ["Stage", "Opus Approach", "GANYIQ Approach", "Key Difference", "Impact"],
    ("0. Video Ingestion", "Accept YouTube URL or upload. Genre classification (auto or manual). Duration limits: 4hr max.", "Same approach. YouTube URL + manual upload.", "Minimal difference. GANYIQ also supports YouTube URL.", "LOW — No competitive advantage here."),
    ("1. Transcription + Diarization", "Whisper-level STT (likely custom finetuned). Speaker diarization (PyAnnote or custom). Word-level timing.", "Deepgram nova-2 → same output quality. Speaker diarization via AV-ASD. Word-level timing.", "Both produce word-level transcript with speaker labels. Opus may use Whisper (open-source advantage? unlikely).", "LOW — Both have good STT."),
    ("2. Candidate Generation", "HYBRID: Scene detection + speaker turns + transcript paragraphs + audio event peaks → candidate windows + optional ClipAnything + keyword filter.", "RULE-BASED: Candidate extraction from transcript → energy + engagement + hook_position scores → top candidates.", "FUNDAMENTAL: Opus uses MULTIPLE signals (visual + audio + transcript + user prompt). GANYIQ uses TRANSCRIPT-ONLY rules.", "HIGH — Opus generates better candidates because it considers more signal types."),
    ("3. Candidate Evaluation", "LLM/ML: Each candidate scored on 3-4 dimensions (Hook, Flow, Value, Trend). Score 0-99 per candidate.", "DETERMINISTIC: Each candidate gets energy*0.4 + engagement*0.35 + hook_position*0.25. No semantic understanding.", "FUNDAMENTAL: Opus UNDERSTANDS content. GANYIQ measures mathematical properties of transcript.", "CRITICAL — This is the biggest quality gap."),
    ("4. Ranking", "Sort: rank (manual override) > score > judgeResult.score. Dedup by task group. pos_rate for normalized position.", "Sort: score descending. No manual rank override. No task grouping. No dedup.", "GANYIQ simpler but misses: manual curation, task grouping, pos_rate normalization.", "MEDIUM — Opus ranking more sophisticated but GANYIQ can replicate."),
    ("5. Filtering", "ClipAnything relevance filter. Keyword filter. Length preferences (clipDurations). iSBonusClip flag for extras.", "Top-N selection by score. No additional filtering.", "GANYIQ has less filtering, which means more low-quality clips may be selected.", "MEDIUM — GANYIQ needs ClipAnything-equivalent and keyword filtering."),
    ("6. Clip Selection Feedback", "Like/dislike per clip. isEdited tracked. ArcQualityFeedback. May retrain from feedback.", "No feedback collection at all.", "CRITICAL — Opus has a learning loop. GANYIQ has none.", "CRITICAL — Without feedback, GANYIQ can't improve over time."),
    ("7. User Override", "rank field allows manual reordering. Results page lets re-prompt ClipAnything. Edit capabilities (trim, extend, fix captions).", "No user override of selection. Accept top candidates.", "GANYIQ gives users no control over selection.", "MEDIUM — User control is product feature, not algorithm."),
]

# ===================================================================
# OPUS_VS_GANYIQ_SELECTION — Head to head comparison
# ===================================================================
OPUS_VS_GANYIQ_SELECTION = [
    ["Scenario", "Opus Would Select", "GANYIQ Would Select", "Why Different", "Winner"],
    ("Podcast: 2 speakers, 1 hour, emotional debate at 23:00, flat intro", "DEBATE CLIP at 23:00. High hook (conflict), high flow (A→B argument), high value (strong opinions), high emotion.", "POSSIBLY the debate IF energy is high. But may also select flat intro if hook_position math works out.", "GANYIQ doesn't understand conflict = high value. Opus scores it higher on ALL 3 dimensions.", "OPUS"),
    ("Educational: Single speaker, 45 min, 3 high-value insights (2:00, 15:00, 32:00), filler between", "3 SEPARATE CLIPS at 2:00, 15:00, 32:00. Each scored independently. Filler correctly rejected.", "MAY pick the same insights if energy peaks there. BUT: may also pick filler if energy calculation errors. NO batch: 1 clip at a time.", "Opus detects content boundaries per insight. GANYIQ may merge or miss them.", "OPUS"),
    ("Vlog: 1 speaker, 20 min, 1 personal story with emotional moment at 12:30, rest is casual talking", "ONE CLIP at 12:30. The emotional moment is the only clip-worthy segment.", "MAY select the emotional moment IF it has highest energy. But may also select casual talking with equal energy.", "GANYIQ's energy scoring may not distinguish between casual talking and emotional moment.", "OPUS (likely)"),
    ("Interview: 2 speakers, 30 min, guest reveals surprising fact at 8:00, host reacts at 8:03", "SINGLE CLIP 7:55-8:30. Surprise fact + host reaction = hook + emotion + dual speaker dynamics.", "MAY split: segment 1 (fact) and segment 2 (reaction separately). Or miss the reaction if no face detection.", "GANYIQ may not connect the fact + reaction as ONE clip. Opus keeps them together as narrative unit.", "OPUS"),
    ("Comedy: 2 speakers, 60 min, joke setup at 42:00, punchline at 42:15, crowd laughter 42:20", "CLIP at 42:00-42:30. Setup + punchline + laughter. Highest emotional peak.", "MAY select the same if energy peaks during laughter. But setup alone may not score high enough.", "Similar for strong audio events. Both detect laughter.", "TIE"),
    ("Business: 1 speaker, 15 min, reveals specific number at 5:00, key framework at 7:00, no laughter", "2 CLIPS: Number reveal (5:00) + Framework explanation (7:00). Value dimension identifies both as valuable.", "MAY pick neither if energy remains low-moderate. Numbers without vocal emphasis don't score high.", "GANYIQ misses authoritative content without vocal energy. Opus identifies it via semantic understanding.", "OPUS"),
    ("Gaming: Gameplay + commentary, key moment at 3:00 with excitement", "GAME LAYOUT clip at 3:00. Game layout unique to Opus. Excitement = peak detection.", "May select via energy peak but outputs in standard layout, not Game layout.", "Layout difference, not selection difference.", "OPUS (layout)"),
]

# ===================================================================
# OPUS_REAL_MOAT_ANALYSIS
# ===================================================================
OPUS_REAL_MOAT = [
    ["Layer", "GANYIQ Can Copy?", "Effort to Copy", "Impact If Copied", "Is This The Moat?"],
    ("Rendering Engine (WASM NLE)", "YES — but not needed immediately", "6-12 months for production quality", "Better preview, faster iteration", "NO — User doesn't see render engine. Output quality is what matters."),
    ("Social Publishing Pipeline", "YES — product feature", "4-8 weeks (API integration)", "Better workflow, user retention", "NO — Can be built. Not a moat."),
    ("Layout Variety (8 layouts)", "YES — partially built already", "4-8 weeks (enable/refine existing)", "More visual variety", "NO — GANYIQ already has most layouts in code, just disabled."),
    ("Subtitle/Brand Templates", "YES — partially built already", "4-8 weeks (template system)", "Better captions, brand consistency", "NO — GANYIQ already has template architecture."),
    ("Batch Processing", "YES — pipeline refactor", "4-8 weeks (pipeline change)", "10x efficiency, scaling", "NO — Straightforward engineering."),
    ("Candidate Generation (Hybrid)", "YES — but harder", "4-12 weeks (algorithm design)", "Better candidate coverage", "POSSIBLY — Important but replicable."),
    ("Scoring Engine (AI 3-dim)", "YES — but requires LLM integration", "8-16 weeks (LLM eval pipeline)", "Single biggest quality improvement", "STRONG CANDIDATE — This is where user-visible quality comes from."),
    ("Feedback Loop + Retraining", "YES — but requires user base", "8-16 weeks (tracking + pipeline)", "Continuous improvement over time", "YES — This compounds over time. Without it, GANYIQ stays static while Opus improves."),
    ("Training Data / Model", "NO — cannot buy", "12+ months (data accumulation)", "The moat that widens over time", "YES — This is the UNCOPYABLE moat. Opus has 5M+ users generating training data."),
    ("Brand + Distribution", "NO — cannot buy", "Years (brand building)", "User acquisition, trust", "YES — Opus brand is established. GANYIQ must build from scratch."),
]

# ===================================================================
# FINAL ANSWER — 3 most valuable capabilities
# ===================================================================
TOP_3_CAPABILITIES = [
    ["Rank", "Capability", "Evidence", "Confidence", "Why #1/2/3"],
    ["#1", "AI Scoring Engine (Hook + Flow + Value + Trend evaluation)", 
     "Official: Virality Score docs. Bundle: 4 scoring dimensions. Feedback loop. Multiple evidence sources converge on this being core value.",
     "HIGH (85%)", "This is what users experience. 'Good clips' = accurate scoring. All other features are supporting cast."],
    ["#2", "Training Data + Feedback Loop (5M+ users generating labeled data)",
     "Bundle: Feedback endpoint. Like/dislike per clip. ArcQualityFeedback feature. isEdited tracking. 5M users from founder talk.",
     "HIGH (80%)", "The moat that widens over time. More data → better model → better clips → more users → more data. GANYIQ starts at zero."],
    ["#3", "Hybrid Candidate Generation (scene + speaker + transcript + audio + prompt)",
     "Bundle: Multiple signal types. ClipAnything with 6 indexed categories. B-roll timeline. Scene/track-based editing script.",
     "MEDIUM (65%)", "Opus considers more signal types than GANYIQ when deciding WHAT could be a clip. This is foundational to #1."],
]

write_sheet("OPUS_GROUND_TRUTH", ["Video ID", "Source", "Category", "Title", "Duration", "Notes"], 
    [("tEXaoozFRes", "YouTube", "Review", "Opus Clip AI review", "~10 min", "Independent review of OpusClip with output examples"),
     ("M52WtXiJQ_s", "YouTube", "Marketing", "OpusClip example compilation", "~5 min", "OpusClip marketing showing multiple output styles"),
     ("b4SSx5f5Q-Q", "YouTube", "Example", "Opus.pro sample clip", "~2 min", "Direct output example from opus.pro"),
     ("5Guxsn7uRd8", "YouTube", "Example", "Opus Clip sample output", "~3 min", "Short-form clip example generated by Opus"),
     ("z9GcDHRMbUk", "YouTube", "Demo", "AI Clip demonstration", "~5 min", "Demo of AI clip generation capabilities"),
     ("reG0NOkFMYs", "YouTube", "Demo", "OpusClip walkthrough", "~8 min", "Product walkthrough showing clip selection"),
     ("wuz2O12wL3g", "YouTube", "User Output", "Made with Opus Clip #shorts", "~1 min", "Real user output from OpusClip"),
     ("nRUDmvIbJkE", "YouTube", "User Output", "Made with OpusClip", "~1 min", "Real user output"),
     ("_jxvx9tzERQ", "YouTube", "User Output", "Opus generated clip", "~1 min", "User-generated clip with Opus"),
     ("ClVGeq5gLgo", "YouTube", "User Output", "OpusClip short form", "~1 min", "Short form content from Opus"),
     ("RQFYS6y4_rs", "YouTube", "User Output", "Opus output example", "~1 min", "Example Opus output"),
     ("LUXaUg7XBR8", "YouTube", "User Output", "Opus clip output", "~1 min", "Opus clip on shorts"),
     ("e9qARr6EIDc", "YouTube", "Review", "Opus AI tool review", "~8 min", "Review with side-by-side comparisons"),
     ("oN27Tj4c7wA", "YouTube", "Demo", "How to use OpusClip", "~6 min", "Tutorial showing workflow"),
     ("Jn1ZnpLuUhw", "YouTube", "Demo", "Opus Clip tutorial", "~7 min", "Complete workflow tutorial")])

write_sheet("OPUS_SELECTION_PATTERNS", OPUS_SELECTION_PATTERNS[0], OPUS_SELECTION_PATTERNS[1:])
write_sheet("OPUS_REJECTION_PATTERNS", OPUS_REJECTION_PATTERNS[0], OPUS_REJECTION_PATTERNS[1:])
write_sheet("OPUS_DECISION_SIMULATION", OPUS_DECISION_SIMULATION[0], OPUS_DECISION_SIMULATION[1:])
write_sheet("OPUS_VS_GANYIQ_SELECTION", OPUS_VS_GANYIQ_SELECTION[0], OPUS_VS_GANYIQ_SELECTION[1:])
write_sheet("OPUS_REAL_MOAT_ANALYSIS", OPUS_REAL_MOAT[0], OPUS_REAL_MOAT[1:])

wb.save(PATH)
print(f"\n✅ Head to head forensics saved to: {PATH}")
print(f"Total sheets: {len(wb.sheetnames)}")
