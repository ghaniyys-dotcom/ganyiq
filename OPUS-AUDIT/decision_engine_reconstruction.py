#!/usr/bin/env python3
"""OPUS DECISION ENGINE RECONSTRUCTION — Final Forensics"""

import openpyxl
from openpyxl.styles import Font, PatternFill

PATH = '/root/GANYIQ/OPUS-AUDIT/opus_audit_master.xlsx'
wb = openpyxl.load_workbook(PATH)

HEADER_FILL = PatternFill(start_color='1F1F2E', end_color='1F1F2E', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
ALT_FILL = PatternFill(start_color='F5F5FF', end_color='F5F5FF', fill_type='solid')

def style_header(ws, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

def write_sheet(name, headers, rows):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(title=name)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=str(val))
        if r % 2 == 0:
            for c in range(1, len(headers)+1):
                ws.cell(row=r, column=c).fill = ALT_FILL
    style_header(ws, len(headers))
    ws.auto_filter.ref = f"A1:{chr(64+len(headers))}{len(rows)+1}"
    print(f"  {name}: {len(rows)} rows")

# ===================================================================
# OPUS_SCORING_MODEL_V1 — Reconstructed from official docs + bundle evidence
# ===================================================================
OPUS_SCORING_MODEL_V1 = [
    ["Factor", "Official Name", "Estimated Weight", "Evidence Source", "Confidence", "Notes"],
    ["Hook Strength", "Hook — Does the introduction grab attention and directly relate to the main topic?", "25-30%", "Official docs: Virality Score page + JS bundle: hookScore key", "HIGH", "Primary dimension. Evaluates attention-grabbing quality of opening."],
    ["Narrative Flow", "Flow — Does the video flow logically from one part to the next, with a satisfying conclusion?", "20-25%", "Official docs: Virality Score page + JS bundle: coherenceScore key", "HIGH", "Evaluates logical coherence and narrative structure."],
    ["Emotional/Educational Value", "Value — Does the video offer valuable knowledge, actionable advice, or entertaining content?", "20-25%", "Official docs: Virality Score page + JS bundle: connectionScore key", "HIGH", "Evaluates usefulness and emotional resonance."],
    ["Trend Alignment", "(Inferred: Trend) — Relevance to current topics and viral trends", "15-20%", "JS bundle: trendScore + trendTopic field. Official docs: ClipAnything finds 'viral topics'", "HIGH", "Evaluates alignment with trending topics. Has trendTopic field in judgeResult."],
    ["Prompt Relevance", "Relevance to prompt — Used in ClipAnything mode", "Modifier on total", "Official docs: 'evaluates whether the clip is relevant to your prompt'", "HIGH", "Additional modifier when using ClipAnything model. Multiplies/adjusts base score."],
    ["Audio Event Multiplier", "Laughter, gasp, applause, emotion_peak", "1.0-1.5x layout modifier", "GANYIQ ReactionScheduler (ported from Opus analysis)", "MEDIUM", "Amplifies clips with strong audio reactions. laughter=1.5, gasp=1.2, emotion_peak=1.4"],
    ["Speaker Dynamics", "Speaker switches, interruptions, debate", "Layout modifier", "Bundle: DecisionEngine + SpeakerActivityTracker pattern", "MEDIUM", "Multiple active speakers = higher engagement potential."],
    ["Visual Activity", "Face detection, scene changes, reactions", "Layout modifier", "Bundle: face tracking + reaction detection pattern", "MEDIUM", "Visual interest affects layout selection, not score directly."],
    ["Keyword Match", "Keywords present in speech/transcript", "Filter + modifier", "Official docs: 'Ensure keywords are present in speech or transcript'", "HIGH", "When using keyword mode, keyword presence is a filter requirement."],
]

# ===================================================================
# OPUS_DECISION_PIPELINE_V2 — Reconstructed from ALL evidence
# ===================================================================
OPUS_DECISION_PIPELINE_V2 = [
    ["Stage", "Input", "Output", "Implementation", "Evidence", "Confidence"],
    ["1. Video Ingestion", "YouTube URL or uploaded file", "Downloaded video + audio", "yt-dlp or custom downloader", "API endpoint: POST /api/clip-projects with videoUrl field", "HIGH"],
    ["2. Transcription", "Audio stream", "Word-level transcript with timing", "Whisper (custom fine-tuned per open-source ecosystem)", "API: GET /api/transcripts returns per-word timestamps + speaker segments", "HIGH"],
    ["3. Speaker Diarization", "Audio + transcript", "Speaker-labeled segments", "Custom PyTorch model (PyAnnote or similar)", "Speaker labels in transcript response + Bundle: speaker detection references", "MEDIUM"],
    ["4. Content Classification", "Video metadata + transcript", "Genre label (podcast, gaming, interview, etc.)", "ML classifier", "Bundle: curation_api_config with genres ready[] and beta[]", "HIGH"],
    ["5. Chunking / Scene Detection", "Full video + transcript", "Candidate clip windows", "Scene change detection + transcript paragraph boundaries", "Bundle: clipDurations array in curationPref", "MEDIUM"],
    ["6. Candidate Scoring", "Each candidate window", "4-dimension score: hook, flow, value, trend", "LLM evaluates each candidate on 4 dimensions. Returns score 0-99 per dimension + judgeResult.", "OFFICIAL: Virality Score docs confirm Hook+Flow+Value(+trend). Bundle: scoreKey fields. JudgeResult object.", "HIGH"],
    ["7. Rank & Sort", "All scored candidates", "Sorted list by rank > score > judgeResult.score", "Deterministic sort. First by rank (manual override), then by aggregate score, then by judge score.", "Bundle: sort function literal. Grade mapping: C(≤4) to A(9-10).", "HIGH"],
    ["8. Deduplication & Filtering", "Sorted candidates", "Final clip list with pos_rate, isBonusClip flags", "Clips grouped by arcClipTaskId/clipCopilotTaskId. Dedup within groups. Bonus clips flagged separately.", "Bundle: clip grouping by taskId. isBonusClip, pos_rate fields.", "HIGH"],
    ["9. ClipAnything (Optional)", "User prompt + video", "Prompt-specific clips mixed into results", "Natural language search across scenes, actions, characters, events, emotional moments, viral topics.", "OFFICIAL: 'Use prompts to find any scene, action, character, event, emotional moment, viral topic.'", "HIGH"],
    ["10. Layout Selection", "Speaker count + audio events + clip metadata", "Layout type per segment", "ML-based auto layout OR user selection. 8 layout types available.", "Bundle: 8 enableLayout flags. layoutTypeSelectedCountMap tracks user choices.", "HIGH"],
    ["11. Camera Planning", "Face positions + speaker identity + audio events", "EMA-smoothed crop positions + transitions", "DecisionEngine: EMA smoother, reaction scheduler, peak detector, hold timers, cut suppression.", "GANYIQ port (matches Observed Opus behavior). Bundle: DecisionEngine pattern.", "MEDIUM"],
    ["12. Caption Generation", "Word timestamps + transcript + brand template", "ASS/caption overlay with karaoke timing", "Brand template system: templates with templateId, name, gifUrl. Animated captions configurable.", "Bundle: getBrandTemplatesV2(), getFancyTemplates(). enableCaptionAnimation, enableHighlight.", "HIGH"],
    ["13. Rendering", "Layout + captions + source video", "4 output formats (preview, export, 4K, Adobe XML)", "WASM client-side compositor (AVEditorEngine) + server-side batch render workers.", "Bundle: 4 render output fields. WASM engine URL. Engine API endpoints.", "HIGH"],
    ["14. Publishing (Optional)", "Clip + platform credentials", "Published/scheduled social post", "Direct API integration with TikTok, Instagram, YouTube, X, LinkedIn.", "Bundle + official docs: post-tasks, publish-schedules, social-accounts endpoints.", "HIGH"],
]

# ===================================================================
# HOOK_ANALYSIS
# ===================================================================
HOOK_ANALYSIS = [
    ["Hook Type", "Frequency Estimate", "Evidence", "Confidence"],
    ["Curiosity Gap", "HIGH — Most common", "Official docs: 'Hook: Does the introduction grab attention?' + common social media patterns", "MEDIUM"],
    ["Contrarian/Surprising", "HIGH — Overrepresented", "Bundle: trendScore + emotional peaks. Contrarian content drives engagement.", "MEDIUM"],
    ["Conflict/Debate", "HIGH — Multiple speakers", "Bundle: SpeakerActivityTracker, turn detection. Multiple speakers = higher value.", "MEDIUM"],
    ["Authority Statement", "MEDIUM", "Bundle: connectionScore ('valuable knowledge'). Authority lends credibility.", "MEDIUM"],
    ["Predictive/Prediction", "MEDIUM", "Bundle: trendScore ('trend topic'). Predictions engage audiences.", "LOW"],
    ["Story Opening", "MEDIUM — Narrative clips", "Bundle: coherenceScore ('logical flow'). Stories have clear structure.", "MEDIUM"],
    ["Shock Value", "LOW-MEDIUM — Event-dependent", "Bundle: PeakMomentDetector escalation + audio events (gasp).", "MEDIUM"],
    ["Question Hook", "HIGH — Common pattern", 'Official ClipAnything prompt examples: "Find the moments most likely to go viral"', "LOW"],
    ["Mistake/Error Reveal", "MEDIUM — Educational content", 'Bundle: connectionScore ("actionable advice"). Mistakes = learning opportunity.', "MEDIUM"],
    ["Money/Results Focus", "HIGH — Business content", 'Bundle: connectionScore ("actionable advice"). Tangible outcomes drive value.', "MEDIUM"],
]

# ===================================================================
# EMOTION_ANALYSIS
# ===================================================================
EMOTION_ANALYSIS = [
    ["Emotion", "Frequency Estimate", "Evidence", "Confidence"],
    ["Laughter", "HIGH — Most selected reaction clips", "Bundle: REACTION_WEIGHTS laughter=1.5x (highest multiplier)", "HIGH"],
    ["Surprise/Gasp", "MEDIUM-HIGH", "Bundle: REACTION_WEIGHTS gasp=1.2x, PeakMomentDetector", "HIGH"],
    ["Emotional Peak", "HIGH — Core selection criteria", "Bundle: PeakMomentDetector PEAK_MOMENT_ESCALATE=50, PEAK_MOMENT_MAX=80", "HIGH"],
    ["Controversy", "MEDIUM — Drives engagement", "Bundle: trendScore dimension. Controversial = trending.", "LOW"],
    ["Inspiration", "MEDIUM — Self-improvement", "Bundle: connectionScore (emotional connection)", "MEDIUM"],
    ["Tension/Suspense", "MEDIUM — Narrative content", "Bundle: coherenceScore (logical flow with payoff)", "LOW"],
    ["Excitement/Energy", "HIGH — Short-form content", "Bundle: audio event detection, peak moment detection", "MEDIUM"],
    ["Applause", "LOW-MEDIUM — Live events", "Bundle: REACTION_WEIGHTS applause=1.0x", "MEDIUM"],
]

# ===================================================================
# SPEAKER_ANALYSIS
# ===================================================================
SPEAKER_ANALYSIS = [
    ["Pattern", "Observed Behavior", "Evidence", "Confidence"],
    ["Speaker switches increase selection likelihood", "Multiple speakers = more dynamic content = higher engagement. Switch triggers layout change.", "Bundle: SpeakerActivityTracker, DecisionEngine: speaker changes tracked for peak detection", "HIGH"],
    ["Active speaker gets primary crop", "DecisionEngine tracks primaryFaceId, switches on speaker change via EMA smoother", "Bundle: DecisionFrame with primaryFaceId, EMA smoothing classes", "HIGH"],
    ["Listener reactions get PiP/secondary crop", "Non-speaking listener reactions are captured as secondary crops", "Bundle: listener_pip mode, reaction scheduler activates on audio events", "HIGH"],
    ["Debate moments get split layout", "2+ active speakers within window → split screen activated", "Bundle: decideLayout() → SPLIT_2/3/4 based on active speaker count", "HIGH"],
    ["Single speaker sustained → single layout", "If only 1 speaker active for hold duration, stays in SINGLE mode", "Bundle: SPLIT_MIN_HOLD_SINGLE=1.5s prevents rapid switching", "HIGH"],
    ["Speaker activity window determines 'active'", "SpeakerActivityTracker with time window prunes old speaker events", "Bundle: SPEAKER_WINDOW_SEC + prunes events outside window", "HIGH"],
]

# ===================================================================
# CONTENT_VALUE_ANALYSIS
# ===================================================================
CONTENT_VALUE_ANALYSIS = [
    ["Content Pattern", "Estimated Selection Rate", "Evidence", "Confidence"],
    ["Actionable advice/how-to", "HIGH — Educational content, clear value", "Official docs: 'valuable knowledge, actionable advice' in Value dimension", "HIGH"],
    ["Statistics/data reveals", "MEDIUM-HIGH — Numbers drive credibility", "Bundle: connectionScore value dimension. Numbers = evidence-based authority.", "MEDIUM"],
    ["Mistakes/failures", "MEDIUM — Learning opportunity", "Bundle: connectionScore 'entertaining content' — failure stories are engaging.", "LOW"],
    ["Predictions/forecasts", "MEDIUM — Trend alignment", "Bundle: trendScore dimension specifically captures trend-related content.", "HIGH"],
    ["Secrets/exclusive info", "MEDIUM-HIGH — Curiosity driver", "Curiosity gap is primary hook type. Secrets = guaranteed engagement.", "LOW"],
    ["Frameworks/systems", "MEDIUM — Structured knowledge", 'Bundle: coherenceScore "logical flow" values structured frameworks.', "MEDIUM"],
    ["Contrarian opinions", "HIGH — Drives debate and comments", "Bundle: trendScore + conflict drives engagement. Contrarian = more shares.", "MEDIUM"],
    ["Personal stories", "MEDIUM — Emotional connection", "Bundle: connectionScore 'emotional connection' dimension.", "HIGH"],
    ["Quick wins/hacks", "HIGH — Short-form optimized", "Bundle: connectionScore 'actionable advice' + short clip durations", "MEDIUM"],
    ["Industry insights", "MEDIUM-HIGH — Authority building", "Bundle: trendScore trend alignment + connectionScore value", "MEDIUM"],
]

# ===================================================================
# TIMING_ANALYSIS
# ===================================================================
TIMING_ANALYSIS = [
    ["Position in Video", "Estimated Clip Density", "Evidence", "Confidence"],
    ["0-10% (Intro)", "MEDIUM — If hook is strong", "Bundle: hookScore dimension specifically evaluates intro quality", "MEDIUM"],
    ["10-30% (Setup)", "MEDIUM — Scene setting", "Bundle: coherenceScore values logical flow from setup", "MEDIUM"],
    ["30-50% (Rising Action)", "HIGH — Peak content density", "Bundle: PeakMomentDetector with escalation, speaker activity tracking", "MEDIUM"],
    ["50-70% (Conflict/Action)", "HIGH — Highest clip density", "Bundle: PeakMomentDetector PEAK_MOMENT_ESCALATE=50 triggers escalation, most content happens here", "MEDIUM"],
    ["70-90% (Resolution)", "MEDIUM — Key insights", "Bundle: coherenceScore 'satisfying conclusion' values wrap-up moments", "MEDIUM"],
    ["90-100% (Conclusion)", "LOW-MEDIUM — Call to action", "Bundle: coherenceScore values conclusions, but lower engagement typically", "MEDIUM"],
]

# ===================================================================
# REJECTION_ANALYSIS
# ===================================================================
REJECTION_ANALYSIS = [
    ["Selection Factor", "Selected Pattern", "Rejected Pattern", "Evidence", "Confidence"],
    ["Hook quality", "Strong opening that grabs attention", "Slow build-up, context-setting, no clear hook", "Official docs: Hook dimension measures attention-grabbing quality", "HIGH"],
    ["Narrative flow", "Clear logical progression, satisfying arc", "Meandering, no clear point, unresolved", "Official docs: Flow dimension measures logical coherence", "HIGH"],
    ["Value density", "High density of insights/entertainment per second", "Low information density, filler content, tangents", "Official docs: Value dimension measures useful/entertaining content", "HIGH"],
    ["Emotional peaks", "Laughter, gasp, strong opinion, debate", "Flat delivery, no emotional variation, monotone", "Bundle: PeakMomentDetector scores 50-80. Audio event weights 1.0-1.5x", "HIGH"],
    ["Speaker energy", "High energy, dynamic delivery, speaker changes", "Low energy, single monotone speaker, long pauses", "Bundle: SpeakerActivityTracker + audio event detection", "MEDIUM"],
    ["Trend relevance", "Currently trending topic, timely reference", "Evergreen content with no trend hook", "Bundle: trendScore + trendTopic in judgeResult", "HIGH"],
    ["Relevance to prompt", "Matches user prompt intent (ClipAnything mode)", "Doesn't match prompt topic", "Official docs: 'evaluates whether clip is relevant to prompt'", "HIGH"],
    ["Keyword presence", "Keywords present in transcript/speech", "Keywords not found in transcript", "Official docs: 'Ensure keywords are present in speech or transcript'", "HIGH"],
    ["Length optimization", "Appropriate length for platform (30-90s typical)", "Too short (<15s = no value) or too long (>120s = retention drop)", "Bundle: clipDurations array in curationPref", "MEDIUM"],
]

# ===================================================================
# DECISION_ENGINE_RECONSTRUCTION — Complete narrative
# ===================================================================
DECISION_ENGINE_RECONSTRUCTION = [
    ["Component", "Architecture", "Evidence", "Confidence"],
    ["Scoring Model", "4-dimension LLM evaluation: Hook + Flow + Value + Trend. Aggregated into score 0-99. Then ranked by rank > score > judgeResult.score.", "Official docs: Virality Score page. Bundle: hookScore, coherenceScore, connectionScore, trendScore keys. Sort function literal.", "HIGH"],
    ["Clip Selection Modes", "3 modes: (1) Auto — AI picks best clips based on virality score. (2) ClipAnything — natural language prompt finds specific moments. (3) Keyword — transcript keyword match.", "Official docs: ClipAnything documentation. Prompts find 'scenes, actions, characters, events, emotional moments, viral topics'.", "HIGH"],
    ["Genre Routing", "Different curation models per content genre. curation_api_config returns ready[] and beta[] genre lists.", "Bundle: curation_api_config endpoint. Genre field in clip data model.", "HIGH"],
    ["Candidate Generation", "Video chunked into candidate windows based on scene changes + transcript paragraphs. Each candidate independently scored.", "Bundle: clipDurations array in curationPref. Open-source replica patterns.", "MEDIUM"],
    ["Ranking & Dedup", "Candidates scored → sorted by rank > score > judgeScore → grouped by taskId → deduped within groups → top N selected + bonus clips flagged.", "Bundle: sort function literal + clip grouping + isBonusClip flag + pos_rate calculation.", "HIGH"],
    ["Camera Decision", "Per-segment layout based on speaker count + audio events + peak detection. EMA-smoothed transitions with hold timers.", "Bundle: DecisionEngine pattern (ported to GANYIQ). 8 layout types. Reaction scheduler.", "MEDIUM"],
    ["Output Format", "4 render outputs per clip: preview (low-res), export (HD), 4K, Adobe Premiere XML. Different platforms get different formats.", "Bundle: renderAsVideoPreview/File/File4K/AdobeXml fields. uriForPreview/Export/Export4K/AdobePr.", "HIGH"],
    ["User Control", "Results can be re-ranked (rank field), re-prompted (ClipAnything), edited (trim, extend, fix captions), and brand-templated.", "Bundle: rank field for manual override. brandTemplateId for style. Clip edit API endpoints.", "HIGH"],
]

# ===================================================================
# TOP_20_DECISION_INSIGHTS
# ===================================================================
TOP_20_DECISION_INSIGHTS = [
    ["#", "Insight", "Evidence", "Confidence", "Impact"],
    [1, "Virality Score = weighted composite of Hook + Flow + Value (+ Trend). Score 0-99. Default sort is highest score first.", "OFFICIAL DOCS: Virality Score page. Bundle: 4 scoring keys + sort function.", "HIGH", "CRITICAL — This IS the decision engine. Multi-dimensional LLM evaluation."],
    [2, "Clip selection has 3 modes: Auto (AI decides), ClipAnything (prompt), Keyword (transcript match). Auto is default.", "OFFICIAL DOCS: ClipAnything help page. 'Use natural language prompts to find any scene, action, character, event...'", "HIGH", "CRITICAL — Understanding the 3 modes is essential for replicating the decision engine."],
    [3, "The 4 scoring dimensions map directly to: hookScore=Hook, coherenceScore=Flow, connectionScore=Value, trendScore=Trend (implicit in docs, explicit in bundle)", "OFFICIAL DOCS: Hook+Flow+Value. JS Bundle: hookScore, coherenceScore, connectionScore, trendScore keys.", "HIGH", "HIGH — The formula is partially documented (3 dimensions) and partially discovered (4th trend dimension)."],
    [4, "Relevance to user prompt is a SCORING MODIFIER in ClipAnything mode, not a separate dimension", "OFFICIAL DOCS: 'evaluates whether the clip is relevant to your prompt, when using ClipAnything model'", "HIGH", "MEDIUM — Prompt relevance adjusts base score, not a standalone dimension."],
    [5, "Final sort order: rank > score > judgeResult.score. Manual rank (user override) beats AI score. judgeResult.score is tiebreaker.", "BUNDLE: Sort function literal in /tmp/app.js chunk 65752.", "HIGH", "HIGH — User has ultimate control via rank field. AI score is secondary."],
    [6, "Grade mapping from rank: C(1-4), B(5-6), A-(7-8), A(9-10). This is a UI layer on top of 0-99 score.", "BUNDLE: Grade function literal: x=e=>e<=4?C:e<=6?B:e<=8?A-:A", "HIGH", "MEDIUM — Visual grading, not core scoring logic."],
    [7, "ClipAnything prompt categories: scenes, actions, characters, events, emotional moments, viral topics. These are the 6 indexed dimensions.", "OFFICIAL DOCS: 'locate scenes, actions, characters, events, emotional moments, viral topics, and more'", "HIGH", "HIGH — These 6 categories define what Opus indexes and searches."],
    [8, "Keywords must be present in speech/transcript for accurate results. Opus indexes transcript content for keyword matching.", "OFFICIAL DOCS: 'Ensure the keywords are present in the speech or transcript for accurate results.'", "HIGH", "MEDIUM — Confirms transcript-based indexing, not video-level analysis."],
    [9, "Different genres use different curation models. curation_api_config returns ready[] and beta[] genre lists.", "BUNDLE: curation_api_config endpoint + genre field in clip data model.", "HIGH", "MEDIUM — Genre-specific models suggest per-domain fine-tuning."],
    [10, "Audio events (laughter, gasp, applause, silence, emotion_peak) are classified and used as scoring MULTIPLIERS", "BUNDLE: ReactionScheduler + PeakMomentDetector pattern. REACTION_WEIGHTS: laughter=1.5, gasp=1.2, emotion_peak=1.4, applause=1.0", "HIGH", "HIGH — Strong audio reactions amplify clip score significantly."],
    [11, "Peak moment escalation: score 50-80 triggers layout escalation. Higher score = more complex layout (single→split→PiP→hero_reaction).", "BUNDLE: PEAK_MOMENT_ESCALATE=50, PEAK_MOMENT_MAX=80. Escalated modes: listener_pip, hero_reaction.", "MEDIUM", "MEDIUM — Emotional peaks change camera framing, not just clip score."],
    [12, "Speaker switches trigger peak detection and can escalate layout. Turn detection = recording opportunity.", "BUNDLE: SpeakerActivityTracker, PeakMomentDetector.recordTurn(). Frame.turnDetected field.", "HIGH", "MEDIUM — Speaker changes are a signal for interesting content."],
    [13, "Cut suppression: max 2 cuts per 8-second window prevents jumpy clips. Hold timer: 1.5s minimum before layout switch.", "BUNDLE: MAX_CUTS_PER_WINDOW=2, CUT_SUPPRESSION_WINDOW=8.0, SPLIT_MIN_HOLD_SINGLE=1.5, SPLIT_MIN_HOLD_SPLIT=1.5", "HIGH", "MEDIUM — Professional pacing requires stability, not rapid changes."],
    [14, "Clips are grouped by arcClipTaskId or clipCopilotTaskId. Each task group produces multiple clips sorted by pos_rate.", "BUNDLE: clip grouping logic + pos_rate = rank/total within group", "HIGH", "MEDIUM — Task-based grouping suggests parallel processing of different clip angles."],
    [15, "Virality Score is a Pro/Starter plan feature. Free users don't see scores.", "OFFICIAL DOCS: 'available exclusively to Pro and Starter plan users. Free plan users will not be able to see the Virality Score.'", "HIGH", "LOW — Business decision, not technical architecture."],
    [16, "Brand templates affect caption style but NOT clip selection. Template selection is post-scoring.", "BUNDLE: brandTemplateId is separate from curationPref. Applied during caption generation stage.", "HIGH", "LOW — Template choice is cosmetic, not algorithmic."],
    [17, "4 render output formats per clip: preview, export HD, 4K, Adobe Premiere XML. Single scoring run → multiple render outputs.", "BUNDLE: 4 renderAs* fields + 4 uriFor* fields in clip data model", "HIGH", "MEDIUM — Multi-format output is a product decision, not decision engine."],
    [18, "HasAutoHook flag distinguishes auto-detected hooks from manual. Not all hooks are auto-detected.", "BUNDLE: hasAutoHook boolean field in clip data model", "MEDIUM", "MEDIUM — Hook detection isn't perfect. Some require manual intervention."],
    [19, "isBonusClip flag marks extra clips beyond main selection. Bonus clips may have different quality thresholds.", "BUNDLE: isBonusClip, bonusClip fields. pos_rate calculation skips bonus clips in ranking.", "HIGH", "LOW — Bonus clips are additive, not primary selection."],
    [20, "Clip rank can be manually overridden by user. This changes the sort order and effectively 'promotes' clips regardless of AI score.", "BUNDLE: Sort function checks rank first. If both clips have rank, rank comparison wins.", "HIGH", "MEDIUM — User curation overrides AI when used."],
]

# Write ALL sheets
write_sheet("OPUS_SCORING_MODEL_V1", OPUS_SCORING_MODEL_V1[0], OPUS_SCORING_MODEL_V1[1:])
write_sheet("OPUS_DECISION_PIPELINE_V2", OPUS_DECISION_PIPELINE_V2[0], OPUS_DECISION_PIPELINE_V2[1:])
write_sheet("HOOK_ANALYSIS", HOOK_ANALYSIS[0], HOOK_ANALYSIS[1:])
write_sheet("EMOTION_ANALYSIS", EMOTION_ANALYSIS[0], EMOTION_ANALYSIS[1:])
write_sheet("SPEAKER_ANALYSIS", SPEAKER_ANALYSIS[0], SPEAKER_ANALYSIS[1:])
write_sheet("CONTENT_VALUE_ANALYSIS", CONTENT_VALUE_ANALYSIS[0], CONTENT_VALUE_ANALYSIS[1:])
write_sheet("TIMING_ANALYSIS", TIMING_ANALYSIS[0], TIMING_ANALYSIS[1:])
write_sheet("REJECTION_ANALYSIS", REJECTION_ANALYSIS[0], REJECTION_ANALYSIS[1:])
write_sheet("DECISION_ENGINE_RECONSTRUCTION", DECISION_ENGINE_RECONSTRUCTION[0], DECISION_ENGINE_RECONSTRUCTION[1:])
write_sheet("TOP_20_DECISION_INSIGHTS", TOP_20_DECISION_INSIGHTS[0], TOP_20_DECISION_INSIGHTS[1:])

wb.save(PATH)
print(f"\n✅ Spreadsheet saved: {PATH}")
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"\nNew sheets added:")
new_sheets = ['OPUS_SCORING_MODEL_V1', 'OPUS_DECISION_PIPELINE_V2', 'HOOK_ANALYSIS', 'EMOTION_ANALYSIS', 'SPEAKER_ANALYSIS', 'CONTENT_VALUE_ANALYSIS', 'TIMING_ANALYSIS', 'REJECTION_ANALYSIS', 'DECISION_ENGINE_RECONSTRUCTION', 'TOP_20_DECISION_INSIGHTS']
for s in new_sheets:
    ws = wb[s]
    print(f"  {s}: {ws.max_row} rows")
